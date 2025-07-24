from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import Device, KeypadEvent, VoteSession
from .serializers import DeviceSerializer, KeypadEventSerializer, LevelTwoUnitWiseSerializer, VoteSessionSerializer
from .services import remote_service
import logging

logger = logging.getLogger(__name__)

# def dashboard(request):
#     """Main dashboard view"""
#     context = {
#         'devices': Device.objects.all()[:10],
#         'recent_events': KeypadEvent.objects.all()[:20],
#         'active_sessions': VoteSession.objects.filter(status='active'),
#         'service_status': remote_service.get_status()
#     }
#     return render(request, 'remote_handler/dashboard.html', context)

# API Views
class DeviceListView(generics.ListAPIView):
    queryset = Device.objects.all()
    serializer_class = DeviceSerializer

class DeviceDetailView(generics.RetrieveAPIView):
    queryset = Device.objects.all()
    serializer_class = DeviceSerializer
    lookup_field = 'base_id'

class KeypadEventListView(generics.ListAPIView):
    queryset = KeypadEvent.objects.all()
    serializer_class = KeypadEventSerializer

    def get_queryset(self):
        queryset = KeypadEvent.objects.all()
        device_id = self.request.query_params.get('device_id', None)
        if device_id is not None:
            queryset = queryset.filter(device__base_id=device_id)
        return queryset

class VoteSessionListView(generics.ListAPIView):
    queryset = VoteSession.objects.all()
    serializer_class = VoteSessionSerializer

@api_view(['POST'])
def start_service(request):
    """Start the remote data collection service"""
    try:
        if remote_service.start_service():
            return Response({
                'status': 'success',
                'message': 'Remote data service started successfully'
            })
        else:
            return Response({
                'status': 'error',
                'message': 'Failed to start service or service already running'
            }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f"Error starting service: {e}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def stop_service(request):
    """Stop the remote data collection service"""
    try:
        remote_service.stop_service()
        return Response({
            'status': 'success',
            'message': 'Remote data service stopped successfully'
        })
    except Exception as e:
        logger.error(f"Error stopping service: {e}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def service_status(request):
    """Get service status"""
    return Response(remote_service.get_status())

@api_view(['GET'])
def stats(request):
    """Get statistics"""
    stats_data = {
        'total_devices': Device.objects.count(),
        'connected_devices': Device.objects.filter(status='connected').count(),
        'total_events': KeypadEvent.objects.count(),
        'unprocessed_events': KeypadEvent.objects.filter(processed=False).count(),
        'active_vote_sessions': VoteSession.objects.filter(status='active').count(),
        'total_vote_sessions': VoteSession.objects.count(),
    }
    return Response(stats_data)

# Add to remote_handler/views.py

@api_view(['POST'])
def start_service_with_params(request):
    """Start the remote data collection service with custom parameters"""
    try:
        base_id = request.data.get('base_id', 1)
        mode = request.data.get('mode', 'auto')
        
        if remote_service.start_service(base_id, mode):
            return Response({
                'status': 'success',
                'message': f'Remote data service started with base_id={base_id}, mode={mode}'
            })
        else:
            return Response({
                'status': 'error',
                'message': f'Failed to start service. Error: {remote_service.last_error}'
            }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f"Error starting service: {e}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def reconnect_service(request):
    """Reconnect to remote device"""
    try:
        base_id = request.data.get('base_id', 1)
        mode = request.data.get('mode', 'auto')
        
        if remote_service.reconnect(base_id, mode):
            return Response({
                'status': 'success',
                'message': 'Reconnected successfully'
            })
        else:
            return Response({
                'status': 'error',
                'message': f'Reconnection failed. Error: {remote_service.last_error}'
            }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

    #-------------------------------------------------------------------------------------------------------------------------------------------------------
    import subprocess
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
#(2) Views for the User Login Views for the User Login
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
# #Create your views here.
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import login
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError
from .serializers import LoginSerializer

class LoginAPIView(APIView):
    """
    User Login API View
    """
    def post(self, request):
        print(request.data)
        serializer = LoginSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            user = serializer.validated_data['user']
            
            # Log the user in
            login(request, user)
            
            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)
            print(user.role)
            
            # Return user info & tokens
            return Response({
                'message': 'Login successful',
                'access_token': access_token,
                'refresh_token': str(refresh),
                'user': {
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'employeeid': user.employeeid,
                    'role': user.role,
                    'hq': user.hq,
                    'factory': user.factory,
                    'department': user.department,
                    'status': user.status
                }
            }, status=status.HTTP_200_OK)
        
        # Extract error message
        error_message = "Authentication failed"
        if serializer.errors:
            for field, errors in serializer.errors.items():
                if errors:
                    if isinstance(errors, list) and errors:
                        error_message = errors[0]
                    else:
                        error_message = str(errors)
                    break
        
        return Response({
            'error': True,
            'message': error_message
        }, status=status.HTTP_400_BAD_REQUEST)
    





#(1) Views for the User Register

from django.db import IntegrityError
from django.shortcuts import render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import User

class RegisterView(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request):
        user_data = request.data
        serializer = self.serializer_class(data=user_data)

        try:
            serializer.is_valid(raise_exception=True)

            # Check if user already exists by email
            if User.objects.filter(email=user_data.get("email")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"email": "This email is already registered."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if employee ID already exists
            if User.objects.filter(employeeid=user_data.get("employeeid")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"employeeid": "This employee ID is already in use."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Save the user
            serializer.save()

            return Response({
                "message": "User registered successfully!"
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            # Handle specific validation errors
            return Response({
                "message": "Validation failed",
                "errors": e.detail
            }, status=status.HTTP_400_BAD_REQUEST)

        except IntegrityError:
            # Handle database integrity errors (like duplicate entries)
            return Response({
                "message": "Database error",
                "errors": {"detail": "Duplicate entry or constraint violation."}
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            # Handle unexpected errors
            return Response({
                "message": "Unexpected error occurred",
                "errors": {"detail": str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#(3) Views for the User Logout

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LogoutSerializer

class LogoutAPIView(APIView):
    """
    User Logout API View
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)
        if serializer.is_valid():
            refresh_token = serializer.validated_data["refresh_token"]

            try:
                token = RefreshToken(refresh_token)
                token.blacklist()  # Blacklist the refresh token
                return Response({"message": "Logout successful"}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    







from rest_framework import viewsets
from .models import HQ, Factory, Department, Line, Level
from .serializers import HQSerializer, FactorySerializer, DepartmentSerializer, LineSerializer, LevelSerializer

class HQViewSet(viewsets.ModelViewSet):
    queryset = HQ.objects.all()
    serializer_class = HQSerializer

class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer

class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer








from rest_framework import viewsets
from .models import SkillTraining
from .serializers import SkillTrainingSerializer

class SkillTrainingViewSet(viewsets.ModelViewSet):
    queryset = SkillTraining.objects.all()
    serializer_class = SkillTrainingSerializer






from rest_framework import viewsets
from .models import SubTopic
from .serializers import SubTopicSerializer

class SubTopicViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicSerializer








from rest_framework import viewsets
from .models import SubTopicContent
from .serializers import SubTopicContentSerializer

class SubTopicContentViewSet(viewsets.ModelViewSet):
    queryset = SubTopicContent.objects.all()
    serializer_class = SubTopicContentSerializer





from rest_framework import viewsets
from .models import Days
from .serializers import DaysSerializer

class DaysViewSet(viewsets.ModelViewSet):
    queryset = Days.objects.all()
    serializer_class = DaysSerializer







from rest_framework import viewsets
from .models import TrainingContent
from .serializers import TrainingContentSerializer

class TrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer




from rest_framework import viewsets
from .models import Test, Question
from .serializers import TestSerializer, QuestionSerializer



class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer


class TestViewSet(viewsets.ModelViewSet):
    queryset = Test.objects.all()
    serializer_class = TestSerializer


from rest_framework import viewsets
from .models import OperatorTestAssignment
from .serializers import OperatorTestAssignmentSerializer

class OperatorTestAssignmentViewSet(viewsets.ModelViewSet):
    queryset = OperatorTestAssignment.objects.all()
    serializer_class = OperatorTestAssignmentSerializer

from rest_framework import viewsets
from .models import OperatorRemoteAssignment
from .serializers import OperatorRemoteAssignmentSerializer

class OperatorRemoteAssignmentViewSet(viewsets.ModelViewSet):
    queryset = OperatorRemoteAssignment.objects.all()
    serializer_class = OperatorRemoteAssignmentSerializer







# from rest_framework import viewsets
# from .models import OperatorMaster
# from .serializers import OperatorMasterSerializer

# class OperatorMasterViewSet(viewsets.ModelViewSet):
#     queryset = OperatorMaster.objects.all()
#     serializer_class = OperatorMasterSerializer





from rest_framework import viewsets
from .models import OperatorMaster, OperatorSkillLevel
from .serializers import OperatorSkillMasterSerializer, OperatorSkillLevelSerializer

class OperatorSkillMasterViewSet(viewsets.ModelViewSet):
    queryset = OperatorMaster.objects.all()
    serializer_class = OperatorSkillMasterSerializer


class OperatorSkillLevelViewSet(viewsets.ModelViewSet):
    queryset = OperatorSkillLevel.objects.all()
    serializer_class = OperatorSkillLevelSerializer






from rest_framework import viewsets
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class TrainingReportViewSet(viewsets.ModelViewSet):
    queryset = TrainingReport.objects.all().order_by('-month')
    serializer_class = TrainingReportSerializer



from rest_framework import viewsets
from .models import UnifiedDefectReport
from .serializers import UnifiedDefectReportSerializer

class UnifiedDefectReportViewSet(viewsets.ModelViewSet):
    queryset = UnifiedDefectReport.objects.all().order_by('-month')
    serializer_class = UnifiedDefectReportSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TrainingReport
from .serializers import TrainingReportSerializer
from django.db.models import Sum

class TrainingSummaryView(APIView):
    def get(self, request):
        summary = TrainingReport.objects.aggregate(
            new_operators_joined=Sum("new_operators_joined"),
            new_operators_trained=Sum("new_operators_trained"),
            total_trainings_planned=Sum("total_trainings_planned"),
            total_trainings_actual=Sum("total_trainings_actual")
        )
        return Response(summary)





from rest_framework.generics import ListAPIView
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class OperatorsJoinedVsTrainedView(ListAPIView):
    queryset = TrainingReport.objects.all().order_by("month")
    serializer_class = TrainingReportSerializer




from rest_framework.views import APIView
from .models import UnifiedDefectReport
from rest_framework.response import Response

class MSILDefectsView(APIView):
    def get(self, request):
        data = UnifiedDefectReport.objects.filter(category='MSIL').order_by('month')
        return Response([
            {
                'month': d.month,
                'total_defects': d.total_defects,
                'ctq_defects': d.ctq_defects,
            }
            for d in data
        ])





class CTQDefectsAllPlantsView(APIView):
    def get(self, request):
        data = UnifiedDefectReport.objects.filter(category='All Plants').order_by('month')
        return Response([
            {
                'month': d.month,
                'total_defects': d.total_defects,
                'ctq_defects': d.ctq_defects,
            }
            for d in data
        ])




class InternalRejectionView(APIView):
    def get(self, request):
        internal = UnifiedDefectReport.objects.all()
        rejection = internal.aggregate(
            total_internal_rejection=Sum('total_internal_rejection'),
            ctq_internal_rejection=Sum('ctq_internal_rejection')
        )
        return Response(rejection)








from rest_framework import viewsets
from .models import  EmployeeMaster, TrainingActivity
from .serializers import EmployeeMasterSerializer, TrainingActivitySerializer



class EmployeeMasterViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMaster.objects.all()
    serializer_class = EmployeeMasterSerializer


class TrainingActivityViewSet(viewsets.ModelViewSet):
    queryset = TrainingActivity.objects.all()
    serializer_class = TrainingActivitySerializer









from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import OperatorMaster, Employee
from .serializers import EmployeeSerializer


class EmployeeByCodeView(APIView):
    def get(self, request, employee_code):
        try:
            operator = OperatorMaster.objects.get(employee_code=employee_code)
            employee = Employee.objects.get(operator=operator)
            serializer = EmployeeSerializer(employee)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except OperatorMaster.DoesNotExist:
            return Response({"error": "Operator not found"}, status=status.HTTP_404_NOT_FOUND)
        except Employee.DoesNotExist:
            return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)














from rest_framework import viewsets
from .models import (
    LevelTwoSkillTraining,
    LevelTwoSection,
    LevelTwoTopic,
    LevelTwoSubTopic,
    LevelTwoUnit,
    LevelTwoSubUnit,
    LevelTwoTrainingContent,
)
from .serializers import (
    LevelTwoSkillTrainingSerializer,
    LevelTwoSectionSerializer,
    LevelTwoTopicSerializer,
    LevelTwoSubTopicSerializer,
    LevelTwoUnitSerializer,
    LevelTwoSubUnitSerializer,
    LevelTwoTrainingContentSerializer,
)


class LevelTwoSkillTrainingViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoSkillTraining.objects.all()
    serializer_class = LevelTwoSkillTrainingSerializer


class LevelTwoSectionViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoSection.objects.all()
    serializer_class = LevelTwoSectionSerializer


class LevelTwoTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTopic.objects.all()
    serializer_class = LevelTwoTopicSerializer


class LevelTwoSubTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoSubTopic.objects.all()
    serializer_class = LevelTwoSubTopicSerializer


class LevelTwoUnitViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoUnit.objects.all()
    serializer_class = LevelTwoUnitSerializer


class LevelTwoSubUnitViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoSubUnit.objects.all()
    serializer_class = LevelTwoSubUnitSerializer


class LevelTwoTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTrainingContent.objects.all()
    serializer_class = LevelTwoTrainingContentSerializer











from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import viewsets
from .models import TrainingContent
from .serializers import TrainingContentSerializer

class SubtopicWiseTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

    @action(detail=False, methods=['get'], url_path='(?P<id>\d+)')
    def subtopicwise(self, request, id=None):
        contents = TrainingContent.objects.filter(subtopic_content_id=id)
        serializer = self.get_serializer(contents, many=True)
        return Response(serializer.data)



class TrainingContentCreateView(APIView):
    def post(self, request):
        serializer = TrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    





from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelTracking
from .serializers import OperatorLevelTrackingSerializer
from datetime import date

@api_view(['GET'])
def get_today_milestones(request):
    today = date.today()
    milestones = OperatorLevelTracking.objects.all()

    # Filter those where join_date + day == today
    matching = [track for track in milestones if track.milestone_date == today]

    serializer = OperatorLevelTrackingSerializer(matching, many=True)
    return Response(serializer.data)











from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

@api_view(['GET'])
def get_unit_by_id(request, unit_id):
    try:
        unit = LevelTwoUnit.objects.get(id=unit_id)
        serializer = LevelTwoUnitSerializer(unit)
        return Response(serializer.data)
    except LevelTwoUnit.DoesNotExist:
        return Response({'error': 'Unit not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
def get_units_by_subtopic(request, subtopic_id):
    units = LevelTwoUnit.objects.filter(subtopic_id=subtopic_id)
    serializer = LevelTwoUnitSerializer(units, many=True)
    return Response(serializer.data)



@api_view(['GET'])
def get_units_by_topic(request, topic_id):
    units = LevelTwoUnit.objects.filter(topic=topic_id, subtopic__isnull=True)
    serializer = LevelTwoUnitWiseSerializer(units, many=True)
    print(serializer.data)
    return Response(serializer.data)



from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import LevelTwoTrainingContent
from .serializers import LevelTwoTrainingContentSerializer

class LevelTwoTrainingSubunitWiseContentViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTrainingContent.objects.all()
    serializer_class = LevelTwoTrainingContentSerializer

    @action(detail=False, methods=['get'], url_path='by-subunit/(?P<subunit_id>[^/.]+)')
    def get_by_subunit(self, request, subunit_id=None):
        contents = self.queryset.filter(subunit_id=subunit_id)
        serializer = self.get_serializer(contents, many=True)
        return Response(serializer.data)




# from rest_framework import viewsets
# from .models import ARVRTrainingContent
# from .serializers import ARVRTrainingContentSerializer

# class ARVRTrainingContentViewSet(viewsets.ModelViewSet):
#     queryset = ARVRTrainingContent.objects.all()
#     serializer_class = ARVRTrainingContentSerializer





# your_app/views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.mail import send_mail
from django.conf import settings

class SendMailView(APIView):
    def post(self, request):
        to_email = request.data.get('to_email')
        subject = request.data.get('subject')
        message = request.data.get('message')

        if not all([to_email, subject, message]):
            return Response({'error': 'Missing required fields'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [to_email],
                fail_silently=False,
            )
            return Response({'success': 'Email sent successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)























# from rest_framework import viewsets
# from .models import OperatorMaster, OperatorPerformanceEvaluation, TrainingCycle, JudgementCriteria
# from .serializers import (
#     OperatorMasterSerializer,
#     OperatorPerformanceEvaluationSerializer,
#     TrainingCycleSerializer,
#     JudgementCriteriaSerializer
# )

# class OperatorPerformanceEvaluationViewSet(viewsets.ModelViewSet):
#     """ViewSet for managing Operator Performance Evaluations"""
#     queryset = OperatorPerformanceEvaluation.objects.all()
#     serializer_class = OperatorPerformanceEvaluationSerializer

# class TrainingCycleViewSet(viewsets.ModelViewSet):
#     """ViewSet for managing Training Cycles"""
#     queryset = TrainingCycle.objects.all()
#     serializer_class = TrainingCycleSerializer

# class JudgementCriteriaViewSet(viewsets.ModelViewSet):
#     """ViewSet for managing Judgement Criteria"""
#     queryset = JudgementCriteria.objects.all()
#     serializer_class = JudgementCriteriaSerializer

# ================================================================
# PERFORMANCE EVALUATION VIEWSETS - END
# ================================================================


#═══════════════════════════════════════════════════════════════
# OBSERVATION AND MONITORING SYSTEM
#═══════════════════════════════════════════════════════════════

# ================================================================
# OBSERVATION MANAGEMENT VIEWSETS - START
# ================================================================

from rest_framework import viewsets
from .models import (
    Process, ObservationItem, PersonnelObservanceSheet,
    ObservationResult, ObservationFailurePoint, Evaluation
)
from .serializers import (
    ProcessSerializer, ObservationItemSerializer,
    PersonnelObservanceSheetSerializer, ObservationResultSerializer,
    ObservationFailurePointSerializer, EvaluationSerializer
)

class ProcessViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Processes"""
    queryset = Process.objects.all()
    serializer_class = ProcessSerializer

class ObservationItemViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Observation Items"""
    queryset = ObservationItem.objects.all()
    serializer_class = ObservationItemSerializer

class PersonnelObservanceSheetViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Personnel Observance Sheets"""
    queryset = PersonnelObservanceSheet.objects.all()
    serializer_class = PersonnelObservanceSheetSerializer

class ObservationResultViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Observation Results"""
    queryset = ObservationResult.objects.all()
    serializer_class = ObservationResultSerializer

class ObservationFailurePointViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Observation Failure Points"""
    queryset = ObservationFailurePoint.objects.all()
    serializer_class = ObservationFailurePointSerializer

class EvaluationViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Evaluations"""
    queryset = Evaluation.objects.all()
    serializer_class = EvaluationSerializer

# ================================================================
# OBSERVATION MANAGEMENT VIEWSETS - END
# ================================================================


#═══════════════════════════════════════════════════════════════
# RE-TRAINING AND CONFIRMATION SYSTEM
#═══════════════════════════════════════════════════════════════

# ================================================================
# RE-TRAINING CONFIRMATION VIEWSETS - START
# ================================================================

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import ReTrainingConfirmation, Observation
from .serializers import ReTrainingConfirmationSerializer, ObservationSerializer
  # Ensure correct path

class ReTrainingConfirmationViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Re-Training Confirmations"""
    queryset = ReTrainingConfirmation.objects.all()
    serializer_class = ReTrainingConfirmationSerializer

    @action(detail=False, methods=['post'])
    def autofill(self, request):
        """Auto-fill operator data based on name and code"""
        emp_name = request.data.get('name')
        emp_code = request.data.get('employee_code')
        try:
            operator = OperatorMaster.objects.get(name=emp_name, code=emp_code)
            data = {
                'operator_id': operator.id,
                'operator_name': operator.name,
                'operator_code': operator.code,
                'operator_join_date': operator.join_date
            }
            return Response(data)
        except OperatorMaster.DoesNotExist:
            return Response({'error': 'Operator not found'}, status=404)

# ================================================================
# RE-TRAINING CONFIRMATION VIEWSETS - END
# ================================================================


# ================================================================
# TRAINER MANAGEMENT VIEWSETS - START
# ================================================================

from rest_framework import viewsets
from .models import Trainer
from .serializers import TrainerSerializer

class TrainerViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Trainers"""
    queryset = Trainer.objects.all()
    serializer_class = TrainerSerializer

# ================================================================
# TRAINER MANAGEMENT VIEWSETS - END
# ================================================================


# ================================================================
# OBSERVATION VIEWSETS - START
# ================================================================

class ObservationViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Observations"""
    queryset = Observation.objects.all()
    serializer_class = ObservationSerializer

# ================================================================
# OBSERVATION VIEWSETS - END
# ================================================================


#═══════════════════════════════════════════════════════════════
# SKILL MATRIX MANAGEMENT SYSTEM
#═══════════════════════════════════════════════════════════════

# ================================================================
# SKILL MATRIX VIEWSETS - START
# ================================================================

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import Operation, Employee, SkillEntry, SkillMatrixMetadata
from .serializers import (
    OperationSerializer, EmployeeSerializer, SkillEntrySerializer,
    SkillMatrixMetadataSerializer, SkillMatrixSerializer
)

class OperationViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Operations"""
    queryset = Operation.objects.all()
    serializer_class = OperationSerializer

class EmployeeViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Employees"""
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

class SkillEntryViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Skill Entries"""
    queryset = SkillEntry.objects.all()
    serializer_class = SkillEntrySerializer

class SkillMatrixMetadataViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Skill Matrix Metadata"""
    queryset = SkillMatrixMetadata.objects.all()
    serializer_class = SkillMatrixMetadataSerializer

# class SkillMatrixViewSet(viewsets.ViewSet):
#     """ViewSet for managing complete Skill Matrix operations"""
    
#     def create(self, request):
#         """Create a new skill matrix entry"""
#         serializer = SkillMatrixSerializer(data=request.data)
#         if serializer.is_valid():
#             result = serializer.save()
#             return Response({"status": "created", "data": SkillMatrixSerializer(result).data})
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#     def list(self, request):
        # """List all skill matrix data with employee skills and metadata"""
        # employees = Employee.objects.all()
        # data = []
        # for emp in employees:
        #     skills = SkillEntry.objects.filter(employee=emp)
        #     metadata = SkillMatrixMetadata.objects.order_by('-updated_on').first()
        #     data.append({
        #         'employee': EmployeeSerializer(emp).data,
        #         'skills': SkillEntrySerializer(skills, many=True).data,
        #         'metadata': SkillMatrixMetadataSerializer(metadata).data if metadata else None
        #     })
        # return Response(data)

# ================================================================
# SKILL MATRIX VIEWSETS - END
# ================================================================

#  Machine allocation IJL

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import Machine
from .serializers import MachineSerializer

class MachineViewSet(viewsets.ModelViewSet):
    queryset = Machine.objects.all()
    serializer_class = MachineSerializer



from rest_framework.response import Response
from django.core.mail import send_mail
from .models import MachineAllocation, MachineAllocationTrackingEmail
from .serializers import MachineAllocationSerializer

class MachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            self.perform_create(serializer)

            allocation = serializer.instance
            subject = "Machine Allocation Approval Request"
            message = (
                f"Machine '{allocation.machine.name}' has been allocated to "
                f"'{allocation.employee.full_name}'.\n\n"
                "Please review and approve this allocation request."
            )
            from_email = None  # Uses DEFAULT_FROM_EMAIL from settings
            recipient_list = list(MachineAllocationTrackingEmail.objects.values_list('email', flat=True))

            if recipient_list:
                send_mail(subject, message, from_email, recipient_list)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        # Only allow updating approval_status
        if set(serializer.validated_data.keys()) != {'approval_status'}:
            return Response(
                {'error': 'Only approval_status can be updated.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        self.perform_update(serializer)
        return Response(serializer.data)


from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Machine, OperationList, OperatorLevel, OperatorMaster
from .serializers import EmployeeNameSerializer, MachineAllocationSerializer

class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationSerializer

    @action(detail=False, methods=['get'], url_path='eligible-employees')
    def eligible_employees(self, request):
        machine_id = request.query_params.get('machine_id')
        if not machine_id:
            return Response({'error': 'machine_id is required'}, status=400)

        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({'error': 'Machine not found'}, status=404)

        operation = OperationList.objects.filter(name__iexact=machine.process).first()
        if not operation:
            return Response({'error': f"Operation '{machine.process}' not found."}, status=404)

        operator_levels = OperatorLevel.objects.filter(operation=operation)
        employee_ids = operator_levels.values_list('employee_id', flat=True).distinct()
        employees = OperatorMaster.objects.filter(id__in=employee_ids)

        serializer = EmployeeNameSerializer(employees, many=True)
        return Response(serializer.data)







#machine allocation viewset
# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from rest_framework.permissions import IsAuthenticated
# from django_filters.rest_framework import DjangoFilterBackend
# from rest_framework import filters
# from django.db import transaction
# from .models import Machine, MachineAllocation, Employee
# from .serializers import (
#     MachineSerializer, MachineAllocationSerializer, EmployeeSerializer,
#     MachineWithAllocationsSerializer, OperatorWithAllocationsSerializer,
#     MachineAllocationCreateSerializer, BulkMachineAllocationSerializer,
#     AutoFillEmployeeSerializer
# )

# class EmployeeViewSet(viewsets.ModelViewSet):
#     """ViewSet for Employee model"""
#     queryset = Employee.objects.all()
#     serializer_class = EmployeeSerializer
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
#     search_fields = ['name', 'emp_code', 'email']
#     ordering_fields = ['name', 'emp_code', 'created_at']
#     ordering = ['-created_at']
    
#     @action(detail=False, methods=['post'])
#     def auto_fill_check(self, request):
#         """Check if employee exists and return their data"""
#         serializer = AutoFillEmployeeSerializer(data=request.data)
#         if serializer.is_valid():
#             employee = serializer.validated_data['employee']
#             employee_serializer = EmployeeSerializer(employee)
#             return Response({
#                 'found': True,
#                 'employee': employee_serializer.data
#             })
#         return Response({
#             'found': False,
#             'errors': serializer.errors
#         }, status=status.HTTP_404_NOT_FOUND)
    
#     @action(detail=True, methods=['get'])
#     def with_allocations(self, request, pk=None):
#         """Get operator with their machine allocations"""
#         operator = self.get_object()
#         serializer = OperatorWithAllocationsSerializer(operator)
#         return Response(serializer.data)

# class MachineViewSet(viewsets.ModelViewSet):
#     """ViewSet for Machine model"""
#     queryset = Machine.objects.all()
#     serializer_class = MachineSerializer
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
#     filterset_fields = ['level']
#     search_fields = ['name']
#     ordering_fields = ['name', 'level', 'created_at']
#     ordering = ['-created_at']
    
    # @action(detail=True, methods=['get'])
    # def with_allocations(self, request, pk=None):
    #     """Get machine with its allocations"""
    #     machine = self.get_object()
    #     serializer = MachineWithAllocationsSerializer(machine)
    #     return Response(serializer.data)
    
    # @action(detail=True, methods=['get'])
    # def available_levels(self, request, pk=None):
    #     """Get available levels for a machine"""
    #     machine = self.get_object()
    #     allocated_levels = machine.allocations.values_list('level', flat=True)
    #     available_levels = []
        
    #     for i in range(1, machine.level + 1):
    #         if i not in allocated_levels:
    #             available_levels.append({
    #                 'level': i,
    #                 'display': f'Level {i}'
    #             })
        
    #     return Response({
    #         'machine': machine.name,
    #         'max_level': machine.level,
    #         'available_levels': available_levels
    #     })

# class MachineAllocationViewSet(viewsets.ModelViewSet):
#     """ViewSet for MachineAllocation model"""
#     queryset = MachineAllocation.objects.select_related('machine', 'employee').all()
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
#     filterset_fields = ['machine', 'employee', 'level']
#     search_fields = ['machine__name', 'employee__name', 'employee__emp_code']
#     ordering_fields = ['created_at', 'level']
#     ordering = ['-created_at']
    
#     def get_serializer_class(self):
#         if self.action == 'create':
#             return MachineAllocationCreateSerializer
#         return MachineAllocationSerializer
    
    # @action(detail=False, methods=['post'])
    # def bulk_create(self, request):
    #     """Create multiple allocations at once"""
    #     serializer = BulkMachineAllocationSerializer(data=request.data)
    #     if serializer.is_valid():
    #         try:
    #             with transaction.atomic():
    #                 result = serializer.save()
    #                 return Response({
    #                     'success': True,
    #                     'created_count': len(result['machine_allocations']),
    #                     'allocations': MachineAllocationSerializer(
    #                         result['machine_allocations'], many=True
    #                     ).data
    #                 }, status=status.HTTP_201_CREATED)
    #         except Exception as e:
    #             return Response({
    #                 'success': False,
    #                 'error': str(e)
    #             }, status=status.HTTP_400_BAD_REQUEST)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # @action(detail=False, methods=['get'])
    # def by_machine(self, request):
    #     """Get allocations grouped by machine"""
    #     machines = Machine.objects.prefetch_related('allocations__employee').all()
    #     result = []
        
    #     for machine in machines:
    #         allocations = machine.allocations.all()
    #         result.append({
    #             'machine': MachineSerializer(machine).data,
    #             'allocations': MachineAllocationSerializer(allocations, many=True).data
    #         })
        
    #     return Response(result)
    
    # @action(detail=False, methods=['get'])
    # def by_employee(self, request):
    #     """Get allocations grouped by employee"""
    #     employees = Employee.objects.prefetch_related('machine_allocations__machine').all()
    #     result = []
        
    #     for employee in employees:
    #         allocations = employee.machine_allocations.all()
    #         if allocations.exists():
    #             result.append({
    #                 'employee': EmployeeSerializer(employee).data,
    #                 'allocations': MachineAllocationSerializer(allocations, many=True).data
    #             })
        
    #     return Response(result)
    
    # @action(detail=False, methods=['post'])
    # def auto_fill_allocation(self, request):
    #     """Create allocation with auto-filled employee data"""
    #     employee_name = request.data.get('employee_name')
    #     employee_code = request.data.get('employee_code')
        
    #     if not employee_name or not employee_code:
    #         return Response({
    #             'error': 'employee_name and employee_code are required'
    #         }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create a copy of request data with auto-fill fields
#         data = request.data.copy()
#         serializer = MachineAllocationCreateSerializer(data=data)
        
#         if serializer.is_valid():
#             allocation = serializer.save()
#             return Response(
#                 MachineAllocationSerializer(allocation).data,
#                 status=status.HTTP_201_CREATED
#             )
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class MachineAllocationBulkViewSet(viewsets.ViewSet):
#     """ViewSet for bulk operations on all models"""
#     permission_classes = [IsAuthenticated]
    
#     @action(detail=False, methods=['post'])
#     def create_complete_setup(self, request):
#         """Create machines, operators, and allocations in one go"""
#         data = request.data
        
#         try:
#             with transaction.atomic():
#                 result = {
#                     'operators': [],
#                     'machines': [],
#                     'allocations': []
#                 }
                
                # # Create operators if provided
                # if 'operators' in data:
                #     for operator_data in data['operators']:
                #         operator_serializer = EmployeeSerializer(data=operator_data)
                #         if operator_serializer.is_valid():
                #             operator = operator_serializer.save()
                #             result['operators'].append(operator_serializer.data)
                #         else:
                #             return Response({
                #                 'error': 'Invalid operator data',
                #                 'details': operator_serializer.errors
                #             }, status=status.HTTP_400_BAD_REQUEST)
                
                # # Create machines if provided
                # if 'machines' in data:
                #     for machine_data in data['machines']:
                #         machine_serializer = MachineSerializer(data=machine_data)
                #         if machine_serializer.is_valid():
                #             machine = machine_serializer.save()
                #             result['machines'].append(machine_serializer.data)
                #         else:
                #             return Response({
                #                 'error': 'Invalid machine data',
                #                 'details': machine_serializer.errors
                #             }, status=status.HTTP_400_BAD_REQUEST)
                
                # # Create allocations if provided
                # if 'allocations' in data:
                #     for allocation_data in data['allocations']:
                #         allocation_serializer = MachineAllocationCreateSerializer(data=allocation_data)
                #         if allocation_serializer.is_valid():
                #             allocation = allocation_serializer.save()
                #             result['allocations'].append(
                #                 MachineAllocationSerializer(allocation).data
                #             )
                #         else:
                #             return Response({
                #                 'error': 'Invalid allocation data',
                #                 'details': allocation_serializer.errors
                #             }, status=status.HTTP_400_BAD_REQUEST)
                
                # return Response({
                #     'success': True,
                #     'result': result
                # }, status=status.HTTP_201_CREATED)
                
    #     except Exception as e:
    #         return Response({
    #             'success': False,
    #             'error': str(e)
    #         }, status=status.HTTP_400_BAD_REQUEST)
    
    # @action(detail=False, methods=['get'])
    # def get_complete_data(self, request):
    #     """Get all data with nested relationships"""
    #     machines = Machine.objects.prefetch_related('allocations__employee').all()
    #     operators = Employee.objects.prefetch_related('machine_allocations__machine').all()
        
    #     return Response({
    #         'machines': MachineWithAllocationsSerializer(machines, many=True).data,
    #         'operators': OperatorWithAllocationsSerializer(operators, many=True).data
    #     })

#═══════════════════════════════════════════════════════════════
# END OF VIEWS FILE
#═══════════════════════════════════════════════════════════════













#IJL SKILL MATRIX VIEWS

from rest_framework import viewsets, generics, status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import (
    SkillMatrix, Section, OperationList, OperatorMaster,
    MultiSkilling, MonthlySkill
)
from .serializers import (
    SkillMatrixSerializer, SectionSerializer, OperationListSerializer, 
    OperatorMasterSerializer, MultiSkillingSerializer, MonthlySkillSerializer
)

class SkillMatrixViewSet(viewsets.ModelViewSet):
    queryset = SkillMatrix.objects.all()
    serializer_class = SkillMatrixSerializer

class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.all()
    serializer_class = SectionSerializer

class OperationListViewSet(viewsets.ModelViewSet):
    queryset = OperationList.objects.all()
    serializer_class = OperationListSerializer

class OperatorMasterViewSet(viewsets.ModelViewSet):
    queryset = OperatorMaster.objects.all()
    serializer_class = OperatorMasterSerializer

class MultiSkillingViewSet(viewsets.ModelViewSet):
    queryset = MultiSkilling.objects.all()
    serializer_class = MultiSkillingSerializer

    def perform_create(self, serializer):
        instance = serializer.save()

        # Create MonthlySkill entry
        try:
            monthly_skill = MonthlySkill.objects.create(multiskilling=instance)
            print(f"✅ Created MonthlySkill for {instance.employee.full_name} - {instance.operation.name}")
        except Exception as e:
            print(f"⚠️  Error creating MonthlySkill: {str(e)}")

        # Create OperatorLevel entry with default level 0 when skill is assigned
        try:
            # Check if OperatorLevel already exists
            existing_level = OperatorLevel.objects.filter(
                employee=instance.employee,
                operation=instance.operation,
                skill_matrix=instance.department
            ).first()

            if not existing_level:
                # Create new OperatorLevel with default level 0
                OperatorLevel.objects.create(
                    employee=instance.employee,
                    operation=instance.operation,
                    skill_matrix=instance.department,
                    level=0  # Default level
                )
                print(f"✅ Created OperatorLevel (Level 0) for {instance.employee.full_name} - {instance.operation.name}")
            else:
                print(f"ℹ️  OperatorLevel already exists for {instance.employee.full_name} - {instance.operation.name}")

        except Exception as e:
            print(f"⚠️  Error creating OperatorLevel: {str(e)}")
            # Don't fail the MultiSkilling creation if OperatorLevel creation fails

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import MultiSkilling, OperatorMaster
from .serializers import MultiSkillingSerializer
from django.utils.timezone import now

class ScheduledEmployeesSkillsView(APIView):
    def get(self, request):
        today = now().date()

        # Fetch all MultiSkilling objects
        multiskills = MultiSkilling.objects.select_related(
            'employee', 'department', 'section', 'operation'
        ).all()

        # Group by employee: create a dictionary mapping employee IDs to a list of skill data
        employee_skills_map = {}

        for skill in multiskills:
            employee_id = skill.employee.id
            if employee_id not in employee_skills_map:
                employee_skills_map[employee_id] = {
                    'employee': employee_id,
                    'employee_name': skill.employee.full_name,
                    'skills': [],
                }
            serializer = MultiSkillingSerializer(skill)
            skill_data = serializer.data

            # Add each skill to employee's skill list
            employee_skills_map[employee_id]['skills'].append(skill_data)

        # Convert dict values to list for response
        response_data = list(employee_skills_map.values())

        return Response(response_data)

class MonthlySkillViewSet(viewsets.ModelViewSet):
    queryset = MonthlySkill.objects.all()
    serializer_class = MonthlySkillSerializer

# Department listing
@api_view(['GET'])
def get_unique_departments(request):
    departments = SkillMatrix.objects.values_list('department', flat=True).distinct()
    return Response(departments)

# Employees with active skills (example filter view)
@api_view(['GET'])
def employees_with_active_skills(request):
    name_query = request.GET.get('name', '')
    operators = OperatorMaster.objects.filter(full_name__icontains=name_query)
    result = []
    for operator in operators:
        active_skills = MultiSkilling.objects.filter(employee=operator, status='active')
        skills = [
            {
                "skill_level": skill.skill_level,
                "date": skill.date,
                "notes": skill.notes,
                "status": skill.status,
                "department": skill.department.department if skill.department else None,
                "section": skill.section.name if skill.section else None,
                "operation": skill.operation.name if skill.operation else None,
            }
            for skill in active_skills
        ]
        result.append({
            "employee_id": operator.id,
            "employee_code": operator.employee_code,
            "full_name": operator.full_name,
            "designation": operator.designation,
            "department": operator.department,
            "department_code": operator.department_code,
            "date_of_join": operator.date_of_join,
            "employee_pattern_category": operator.employment_pattern_category,
            "skills": skills
        })
    return Response(result, status=status.HTTP_200_OK)



@api_view(['PUT'])
def update_skillmatrix_by_id(request, id):
    try:
        skill_matrix = SkillMatrix.objects.get(id=id)
    except SkillMatrix.DoesNotExist:
        return Response({"error": "Record not found"}, status=404)

    serializer = SkillMatrixSerializer(skill_matrix, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=400)





from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevel
from .serializers import FirstOperatorLevelSerializer

@api_view(['GET'])
def get_operator_levels_by_department(request, department):
    data = OperatorLevel.objects.filter(skill_matrix__department__iexact=department)
    serializer = FirstOperatorLevelSerializer(data, many=True)
    return Response(serializer.data)


from rest_framework import viewsets
from .models import OperatorLevel
from .serializers import ( OperatorLevelSerializer )

class OperatorLevelViewSet(viewsets.ModelViewSet):
    queryset = OperatorLevel.objects.all()
    serializer_class = OperatorLevelSerializer




# views.py
# from rest_framework.decorators import api_view
# from rest_framework.response import Response
# from .models import OperatorLevel
# from .serializers import FirstOperatorLevelSerializer

# @api_view(['GET'])
# def get_operator_levels_by_department(request, department):
#     data = OperatorLevel.objects.filter(skill_matrix__shop_department=department)
#     serializer = FirstOperatorLevelSerializer(data, many=True)
#     return Response(serializer.data)













# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from tablib import Dataset
from .resources import OperatorMasterResource

class OperatorExcelUploadAPIView(APIView):
    def post(self, request, format=None):
        excel_file = request.FILES.get('file')

        if not excel_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            dataset = Dataset()
            imported_data = dataset.load(excel_file.read(), format='xlsx')

            print(f"Rows in Excel (excluding header): {len(imported_data)}")
            if len(imported_data) == 0:
                return Response({"error": "Excel file is empty or headers do not match."}, status=status.HTTP_400_BAD_REQUEST)

            # Print sample row to verify data parsing
            print("Sample row:", imported_data.dict[0])

            resource = OperatorMasterResource()

            # Dry run to validate data without saving
            dry_run_result = resource.import_data(imported_data, dry_run=True)
            if dry_run_result.has_errors():
                return Response({
                    "error": "Data has validation errors.",
                    "details": [str(err) for err in dry_run_result.row_errors()]
                }, status=status.HTTP_400_BAD_REQUEST)

            # Actual import (dry_run=False saves the data)
            import_result = resource.import_data(imported_data, dry_run=False)
            print("Import result totals:", import_result.totals)

            if import_result.has_errors():
                return Response({
                    "error": "Errors occurred during import.",
                    "details": [str(err) for err in import_result.row_errors()]
                }, status=status.HTTP_400_BAD_REQUEST)

            new_records = import_result.totals.get('new', 0)
            updated_records = import_result.totals.get('update', 0)
            skipped_records = import_result.totals.get('skip', 0)
            total_rows = import_result.totals.get('total_rows', len(imported_data))

            if new_records == 0 and updated_records == 0:
                return Response({
                    "message": "No new or updated records were imported.",
                    "details": {
                        "total_rows": total_rows,
                        "new_records": new_records,
                        "updated_records": updated_records,
                        "skipped_records": skipped_records,
                    }
                }, status=status.HTTP_200_OK)

            return Response({
                "message": "Data imported successfully.",
                "details": {
                    "total_rows": total_rows,
                    "new_records": new_records,
                    "updated_records": updated_records,
                    "skipped_records": skipped_records,
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)






#AR-VR IJL


from rest_framework import viewsets
from .models import ARVRTrainingContent
from .serializers import ARVRTrainingContentSerializer

class ARVRTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ARVRTrainingContent.objects.all()
    serializer_class = ARVRTrainingContentSerializer


# Notification IJL


from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelTracking, OperatorMaster
from .serializers import OperatorLevelTrackingSerializer
from datetime import date, timedelta

@api_view(['GET'])
def get_today_milestones(request):
    today = date.today()
    milestone_rules = OperatorLevelTracking.objects.all()
    milestone_data = []

    for rule in milestone_rules:
        expected_join_date = today - timedelta(days=rule.day)
        matched_employees = OperatorMaster.objects.filter(joining_date=expected_join_date)

        for employee in matched_employees:
            milestone_data.append({
                "operator_name": employee.full_name,
                "level_name": rule.level.name,
                "day": rule.day,
                "milestone_date": today,
                "message": f"{employee.full_name} has completed milestone: {rule.level.name} on Day {rule.day}"
            })

    return Response({
        "date": str(today),
        "milestones": milestone_data
    })






from datetime import date, timedelta
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelEmailTracking, TrackingEmail, OperatorMaster
from .utils import send_milestone_email

@api_view(['GET'])
def get_today_email_milestones(request):
    today = date.today()
    rules = OperatorLevelEmailTracking.objects.all()
    email_logs = []

    for rule in rules:
        milestone_day = rule.day
        expected_date = today - timedelta(days=milestone_day)

        # Get all employees who joined exactly 'day' days ago
        matching_employees = OperatorMaster.objects.filter(date_of_join=expected_date)

        recipient_list = [e.email for e in rule.emails.all() if e.email]

        for employee in matching_employees:
            if recipient_list:
                subject = "Milestone Alert"
                message = f"{employee.full_name} has reached milestone: {rule.level.name} on Day {rule.day}."
                send_milestone_email(subject, message, recipient_list)
                email_logs.append({
                    "employee": employee.full_name,
                    "joined_on": str(employee.date_of_join),
                    "level": rule.level.name,
                    "day": rule.day,
                    "recipients": recipient_list,
                    "status": "Email sent"
                })

    return Response({
        "message": "Milestone emails sent to matching employees.",
        "email_logs": email_logs
    })


#10 cycle ijl

 

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import OperatorMaster
from .serializers import OperatorMasterSerializer

class OperatorMasterViewSet(viewsets.ModelViewSet):
    queryset = OperatorMaster.objects.all()
    serializer_class = OperatorMasterSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee_code']

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
        try:
          employee = OperatorMaster.objects.get(employee_code=employee_code)
          serializer = self.get_serializer(employee)
          return Response(serializer.data)
        except OperatorMaster.DoesNotExist:
          return Response({'error': 'Employee not found'}, status=404)

from collections import defaultdict

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import OperatorPerformanceEvaluation, Tencycletopics, OperatorEvaluation, EvaluationTopicMarks
from .serializers import (
    OperatorPerformanceEvaluationSerializer,
    TencycletopicsSerializer,
    OperatorEvaluationSerializer,
    EvaluationTopicMarksSerializer
)
from .serializers import OperatorPerformanceEvaluationWriteSerializer
from .serializers import EvaluationTopicMarksWriteSerializer

class OperatorPerformanceEvaluationViewSet(viewsets.ModelViewSet):
    queryset = OperatorPerformanceEvaluation.objects.all()

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return OperatorPerformanceEvaluationSerializer  # Read serializer
        return OperatorPerformanceEvaluationWriteSerializer  # Write serializer


class TencycletopicsViewSet(viewsets.ModelViewSet):                  # 10 cycle topics ....
    queryset = Tencycletopics.objects.all()
    serializer_class = TencycletopicsSerializer

class OperatorEvaluationViewSet(viewsets.ModelViewSet):              #day1 to day6 ....
    queryset = OperatorEvaluation.objects.all()
    serializer_class = OperatorEvaluationSerializer


class EvaluationTopicMarksViewSet(viewsets.ModelViewSet):
    queryset = EvaluationTopicMarks.objects.all()

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve', 'by_employee_code']:
            return EvaluationTopicMarksSerializer  # Read (nested)
        return EvaluationTopicMarksWriteSerializer  # Write (IDs)

    

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
    # Step 1: Get the related OperatorPerformanceEvaluation entries
        employee_evals = OperatorPerformanceEvaluation.objects.filter(cc_no__employee_code=employee_code)
        if not employee_evals.exists():
            return Response({'error': 'Employee not found'}, status=404)

    # Step 2: Get all EvaluationTopicMarks for this employee
        evaluations = EvaluationTopicMarks.objects.filter(
        employee__in=employee_evals
    ).select_related('topic_name', 'days', 'employee')

    # Step 3: Serialize data
        serializer = self.get_serializer(evaluations, many=True)

    # Step 4: Group evaluations by day
        days_map = defaultdict(list)
        for eval_obj, eval_data in zip(evaluations, serializer.data):
            days_map[eval_obj.days.day].append(eval_data)

    # Step 5: Score computation
        per_day_results = []
        total_score = 0
        total_possible_score = 0

        for day, topics in days_map.items():
            mark_totals = [0] * 10
            complete = True
            for topic in topics:
                marks = [topic.get(f'mark_{i}', None) for i in range(1, 11)]
                if None in marks:
                    complete = False
                    break
                for i, mark in enumerate(marks):
                    mark_totals[i] += mark if mark is not None else 0

            if complete:
                day_score = sum(mark_totals)
                total_score += day_score
                total_possible_score += 220
                status = "Pass" if day_score >= 0.6 * 220 else "Fail - Retraining Required"
                per_day_results.append({
                "day": day,
                "score": day_score,
                "status": status
            })

        final_status = "Not Evaluated"
        if total_possible_score > 0:
            percentage = (total_score / total_possible_score) * 100
            final_status = "Pass" if percentage >= 60 else "Fail - Retraining Required"

    # Step 6: Optional — include employee basic details
        employee_details = employee_evals.first()
        employee_info = {
        "id": employee_details.id,
        "cc_no": employee_details.cc_no.employee_code,
        "line": employee_details.line,
        "process_name": employee_details.process_name,
        "operation_no": employee_details.operation_no,
        "date_of_retraining_completed": employee_details.date_of_retraining_completed,
        "department": getattr(employee_details.cc_no, 'department', None),
        "prepared_by": employee_details.prepared_by,
        "checked_by": employee_details.checked_by,
        "approved_by": employee_details.approved_by,
        }

        return Response({
        'employee_code': employee_code,
        'employee_details': employee_info,
        'evaluations': serializer.data,
        'per_day_results': per_day_results,
        'total_score': total_score,
        'final_status': final_status
    })
    @action(detail=False, methods=['put'], url_path='update/(?P<employee_code>[^/.]+)')
    def update_by_employee_code(self, request, employee_code=None):
        topic_id = request.data.get("topic_name")
        day_id = request.data.get("days")

        if not topic_id or not day_id:
            return Response({"error": "Both topic_name and days are required in the request body."}, status=400)

        try:
        # Get the performance record using employee_code
            employee_obj = OperatorPerformanceEvaluation.objects.get(cc_no__employee_code=employee_code)
        except OperatorPerformanceEvaluation.DoesNotExist:
            return Response({"error": f"Employee with code {employee_code} not found."}, status=404)

        try:
        # Find the EvaluationTopicMarks record
            instance = EvaluationTopicMarks.objects.get(
            employee=employee_obj,
            topic_name_id=topic_id,
            days_id=day_id
        )
        except EvaluationTopicMarks.DoesNotExist:
            return Response({"error": "Matching evaluation record not found. Nothing to update."}, status=404)

    # Update it using serializer
        serializer = EvaluationTopicMarksWriteSerializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=200)
    
        return Response(serializer.errors,status=400)



#10 cycle for level (new)


from collections import defaultdict
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from .models import (
    OperatorMaster,
    OperatorPerformanceEvaluationLevel,
    TencycletopicsLevel,
    OperatorEvaluationLevel,
    EvaluationTopicMarksLevel
)
from .serializers import (
    OperatorMasterSerializer,
    OperatorPerformanceEvaluationLevelSerializer,
    TencycletopicsLevelSerializer,
    OperatorEvaluationLevelSerializer,
    EvaluationTopicMarksLevelSerializer,
    OperatorPerformanceEvaluationLevelWriteSerializer,
    EvaluationTopicMarksLevelWriteSerializer
)

class OperatorPerformanceEvaluationLevelViewSet(viewsets.ModelViewSet):
    queryset = OperatorPerformanceEvaluationLevel.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['cc_no_l__employee_code']

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return OperatorPerformanceEvaluationLevelSerializer  # Read serializer
        return OperatorPerformanceEvaluationLevelWriteSerializer  # Write serializer

class TencycletopicsLevelViewSet(viewsets.ModelViewSet):
    queryset = TencycletopicsLevel.objects.all()
    serializer_class = TencycletopicsLevelSerializer

class OperatorEvaluationLevelViewSet(viewsets.ModelViewSet):
    queryset = OperatorEvaluationLevel.objects.all()
    serializer_class = OperatorEvaluationLevelSerializer

class EvaluationTopicMarksLevelViewSet(viewsets.ModelViewSet):
    queryset = EvaluationTopicMarksLevel.objects.all()

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve', 'by_employee_code']:
            return EvaluationTopicMarksLevelSerializer  # Read (nested)
        return EvaluationTopicMarksLevelWriteSerializer  # Write (IDs)

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
        # Step 1: Get related OperatorPerformanceEvaluationLevel entries
        employee_evals = OperatorPerformanceEvaluationLevel.objects.filter(cc_no_l__employee_code=employee_code)
        if not employee_evals.exists():
            return Response({'error': 'Employee not found'}, status=404)

        # Step 2: Get all EvaluationTopicMarksLevel for this employee
        evaluations = EvaluationTopicMarksLevel.objects.filter(
            employee_l__in=employee_evals
        ).select_related('topic_name_l', 'days_l', 'employee_l')

        # Step 3: Serialize data
        serializer = self.get_serializer(evaluations, many=True)

        # Step 4: Group evaluations by day
        days_map = defaultdict(list)
        for eval_obj, eval_data in zip(evaluations, serializer.data):
            days_map[eval_obj.days_l.day_l].append(eval_data)

        # Step 5: Score computation
        per_day_results = []
        total_score = 0
        total_possible_score = 0

        for day, topics in days_map.items():
            mark_totals = [0] * 10
            complete = True
            for topic in topics:
                marks = [topic.get(f'mark_{i}', None) for i in range(1, 11)]
                if None in marks:
                    complete = False
                    break
                for i, mark in enumerate(marks):
                    mark_totals[i] += mark if mark is not None else 0

            if complete:
                day_score = sum(mark_totals)
                total_score += day_score
                total_possible_score += 220
                status = "Pass" if day_score >= 0.6 * 220 else "Fail - Retraining Required"
                per_day_results.append({
                    "day": day,
                    "score": day_score,
                    "status": status
                })

        final_status = "Not Evaluated"
        if total_possible_score > 0:
            percentage = (total_score / total_possible_score) * 100
            final_status = "Pass" if percentage >= 60 else "Fail - Retraining Required"

        # Step 6: Optional — include employee basic details
        employee_details = employee_evals.first()
        employee_info = {
            "id": employee_details.id,
            "cc_no": employee_details.cc_no_l.employee_code,
            "line": employee_details.line_l,
            "process_name": employee_details.process_name_l,
            "operation_no": employee_details.operation_no_l,
            "date_of_retraining_completed": employee_details.date_of_retraining_completed_l,
            "department": getattr(employee_details.cc_no_l, 'department', None),
            "prepared_by": employee_details.prepared_by_l,
            "checked_by": employee_details.checked_by_l,
            "approved_by": employee_details.approved_by_l,
        }

        return Response({
            'employee_code': employee_code,
            'employee_details': employee_info,
            'evaluations': serializer.data,
            'per_day_results': per_day_results,
            'total_score': total_score,
            'final_status': final_status
        })

    @action(detail=False, methods=['put'], url_path='update/(?P<employee_code>[^/.]+)')
    def update_by_employee_code(self, request, employee_code=None):
        topic_id = request.data.get("topic_name_l")
        day_id = request.data.get("days_l")

        if not topic_id or not day_id:
            return Response({"error": "Both topic_name_l and days_l are required in the request body."}, status=400)

        try:
            # Get the performance record using employee_code
            employee_obj = OperatorPerformanceEvaluationLevel.objects.get(cc_no_l__employee_code=employee_code)
        except OperatorPerformanceEvaluationLevel.DoesNotExist:
            return Response({"error": f"Employee with code {employee_code} not found."}, status=404)

        try:
            # Find the EvaluationTopicMarksLevel record
            instance = EvaluationTopicMarksLevel.objects.get(
                employee_l=employee_obj,
                topic_name_l_id=topic_id,
                days_l_id=day_id
            )
        except EvaluationTopicMarksLevel.DoesNotExist:
            return Response({"error": "Matching evaluation record not found. Nothing to update."}, status=404)

        # Update it using serializer
        serializer = EvaluationTopicMarksLevelWriteSerializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=200)

        return Response(serializer.errors, status=400)











from rest_framework import viewsets
from .models import SubTopic
from .serializers import  SubTopicDaySerializer

class SubTopicDayViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicDaySerializer



# easy test


# easytest


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from rest_framework import viewsets


class KeyEventCreateView(APIView):
    def post(self, request):
        serializer = KeyEventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Key event saved'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LatestKeyEventView(APIView):
    def get(self, request):
        try:
            latest_event = KeyEvent.objects.latest('timestamp')
            serializer = KeyEventSerializer(latest_event)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyEvent.DoesNotExist:
            return Response({"message": "No key events yet."}, status=status.HTTP_404_NOT_FOUND)

        
# api/views.py
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import ConnectEventSerializer

@api_view(['POST'])
def connect_event_create(request):
    serializer = ConnectEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




@api_view(['POST'])
def vote_event_create(request):
    serializer = VoteEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import Question, EmployeeMaster, Level, Station, TestSession, Score
from .serializers import QuestionSerializer, EmployeeSerializer, ScoreSerializer, SimpleScoreSerializer

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import QuizQuestion
from .serializers import QuizQuestionSerializer


from rest_framework import generics

class QuizQuestionListCreateView(generics.ListCreateAPIView):
    queryset = QuizQuestion.objects.all()
    serializer_class = QuizQuestionSerializer

    def get_queryset(self):
        paper_id = self.request.query_params.get('paper_id')
        if paper_id:
            return self.queryset.filter(question_paper__id=paper_id)
        return self.queryset
    
from rest_framework import generics

class QuizQuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = QuizQuestion.objects.all()
    serializer_class = QuizQuestionSerializer




class QuizQuestionPaperListCreateView(generics.ListCreateAPIView):
    queryset = QuizQuestionPaper.objects.all()
    serializer_class = QuizQuestionPaperSerializer

class QuizQuestionsByPaperView(generics.ListAPIView):
    serializer_class = QuizQuestionSerializer

    def get_queryset(self):
        paper_id = self.kwargs.get('paper_id')
        return QuizQuestion.objects.filter(question_paper_id=paper_id)




class OperatorListCreateView(APIView):
    def get(self, request):
        employees = OperatorMaster.objects.all()
        serializer = OperatorMasterSerializer(employees, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = OperatorMasterSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from django.core.cache import cache
from .serializers import ScoreSerializer

class ScoreListView(APIView):
    def get(self, request):
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(test_name=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)



class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({s.key_id: s.employee.full_name for s in mapping})


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import EmployeeMaster, Station, TestSession, QuizQuestionPaper, Level

class StartTestSessionView(APIView):
    def post(self, request):
        try:
            test_name = request.data.get("test_name")
            assignments = request.data.get("assignments", [])
            question_paper_id = request.data.get("question_paper_id") or request.data.get("paper_id")
            skill_id = request.data.get("skill_id")
            level_id = request.data.get("level_id")

            if not test_name or not assignments:
                return Response(
                    {"error": "Test name and assignments are required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            question_paper = None
            if question_paper_id:
                question_paper = get_object_or_404(QuizQuestionPaper, id=question_paper_id)

            # Get skill and level objects if provided
            skill = None
            level = None

            if skill_id:
                try:
                    skill = Station.objects.get(id=skill_id)
                except Station.DoesNotExist:
                    return Response(
                        {"error": f"Station with id {skill_id} not found. Please check if the station exists."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                except ValueError:
                    return Response(
                        {"error": f"Invalid skill_id format: {skill_id}. Must be a valid integer."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if level_id:
                try:
                    level = Level.objects.get(id=level_id)
                except Level.DoesNotExist:
                    return Response(
                        {"error": f"Level with id {level_id} not found. Please check if the level exists."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                except ValueError:
                    return Response(
                        {"error": f"Invalid level_id format: {level_id}. Must be a valid integer."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # Handle assignments as dictionary (from frontend) or list (legacy)
            if isinstance(assignments, dict):
                # Convert dictionary to list format
                assignment_items = []
                for key_id, assignment_data in assignments.items():
                    assignment_items.append({
                        "key_id": key_id,
                        "employee_id": assignment_data.get("employee_id")
                    })
                assignments = assignment_items

            for item in assignments:
                key_id = item.get("key_id")
                employee_id = item.get("employee_id")

                if not key_id or not employee_id:
                    return Response(
                        {"error": "key_id, employee_id are required in each assignment."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                employee = get_object_or_404(OperatorMaster, id=employee_id)

                # Check if employee has already taken this question paper
                if question_paper:
                    existing_score = Score.objects.filter(
                        employee=employee,
                        test=question_paper
                    ).first()

                    if existing_score:
                        return Response(
                            {"error": f"Employee {employee.full_name} (ID: {employee.id}) has already taken this test. Each employee can only take the same question paper once."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                # Handle existing TestSession - delete if exists to avoid UNIQUE constraint error
                existing_session = TestSession.objects.filter(
                    test_name=test_name,
                    key_id=key_id
                ).first()

                if existing_session:
                    existing_session.delete()
                    print(f"Deleted existing TestSession: {test_name} - {key_id}")

                # Create TestSession with dynamic skill and level
                TestSession.objects.create(
                    test_name=test_name,
                    key_id=key_id,
                    employee=employee,
                    level=level,
                    skill=skill,
                    question_paper=question_paper,
                )

            return Response({"status": "ok"}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, QuizQuestion, Score

class EndTestSessionView(APIView):
    def post(self, request):
        try:
            key_id_to_answers = request.data  # { key_id: [answers] }
            results = []
            test_name = ''
            processed_employees = {}  # Track processed employees with their data

            sessions = TestSession.objects.select_related('employee', 'skill', 'level', 'question_paper').all()

            # If no sessions exist, this might be a duplicate API call after TestSessions were already deleted
            if sessions.count() == 0:
                return Response({'test_name': '', 'results': []}, status=200)

            # First pass: collect unique employee-test combinations and their answers
            for session in sessions:
                key_id = str(session.key_id)
                employee = session.employee
                test_name = session.test_name
                question_paper = session.question_paper

                # Create unique identifier for employee-test combination
                employee_test_key = f"{employee.id}_{test_name}_{question_paper.id if question_paper else 'none'}"

                # Skip if we've already processed this employee for this test
                if employee_test_key in processed_employees:
                    continue

                # Store the employee data for processing
                processed_employees[employee_test_key] = {
                    'session': session,
                    'employee': employee,
                    'test_name': test_name,
                    'question_paper': question_paper,
                    'key_id': key_id,
                    'answers': key_id_to_answers.get(key_id, [])
                }

            # Second pass: process each unique employee-test combination
            for employee_test_key, data in processed_employees.items():
                session = data['session']
                employee = data['employee']
                test_name = data['test_name']
                question_paper = data['question_paper']
                answers = data['answers']

                # Use get_or_create to prevent duplicate Score records at database level
                from django.db import transaction

                try:
                    with transaction.atomic():
                        score, created = Score.objects.get_or_create(
                            employee=employee,
                            test_name=test_name,
                            test=question_paper,
                            defaults={
                                'marks': 0,  # Will be updated below
                                'percentage': 0,  # Will be updated below
                                'passed': False,  # Will be updated below
                                'level': session.level,
                                'skill': session.skill,
                            }
                        )

                        if created:
                            # Calculate score for new record
                            questions = list(QuizQuestion.objects.filter(question_paper=question_paper).order_by('id'))
                            correct_count = 0
                            for i, ans in enumerate(answers):
                                if i < len(questions) and ans == questions[i].correct_index:
                                    correct_count += 1

                            # Calculate percentage with proper rounding
                            total_questions = len(questions)
                            percentage = round((correct_count / total_questions) * 100, 2) if total_questions > 0 else 0
                            passed = percentage >= 80

                            # Update the score with calculated values
                            score.marks = correct_count
                            score.percentage = percentage
                            score.passed = passed
                            score.save()

                            print(f"✅ Created new Score (ID: {score.id}) for {employee.full_name}")
                        else:
                            # Use existing score
                            correct_count = score.marks
                            percentage = score.percentage
                            passed = score.passed
                            total_questions = len(QuizQuestion.objects.filter(question_paper=question_paper))
                            print(f"⚠️ Using existing Score (ID: {score.id}) for {employee.full_name}")

                except Exception as e:
                    print(f"❌ Error creating/getting Score: {str(e)}")
                    continue

                # Automatically update skill matrix if passed
                if passed:
                    skill_matrix_update = score.update_skill_matrix()
                    if skill_matrix_update:
                        print(f"Updated skill matrix for {employee.full_name}: {session.skill.skill if session.skill else 'Unknown'} Level {score.get_level_number()}")

                result_entry = {
                    'name': employee.full_name,
                    'marks': correct_count,
                    'percentage': percentage,
                    'passed': passed,
                    'total_questions': total_questions
                }
                results.append(result_entry)

            TestSession.objects.all().delete()

            return Response({'test_name': test_name, 'results': results}, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)



class PastTestSessionsView(APIView):
    def get(self, request):
        qs = Score.objects.values('test_name').distinct()
        return Response([s['test_name'] for s in qs])


class ScoresByTestView(APIView):
    def get(self, request, name):
        scores = (
            Score.objects
            .filter(test_name=name)
            .select_related('employee', 'level', 'skill')
        )

        questions_count = QuizQuestion.objects.count() or 1

        data = []
        for s in scores:
            data.append({
                'employee_id': s.employee.id,
                'name': s.employee.full_name,
                'marks': s.marks,
                'percentage': round((s.marks / questions_count) * 100, 2),
                'level_name': s.level.get_name_display() if s.level else 'N/A',
                'level_number': s.get_level_number() if s.level else 'N/A',
                'skill': s.skill.skill if s.skill else 'N/A',
                'skill_name': s.skill.skill if s.skill else 'N/A',
                'section': s.employee.department if s.employee.department else 'N/A',
                'passed': s.passed if hasattr(s, 'passed') else (s.marks >= (questions_count * 0.8)),
            })

        return Response(data)


class ResultSummaryAPIView(APIView):
    def get(self, request):
        scores = Score.objects.select_related('employee', 'level', 'skill')
        data = []
        for score in scores:
            percentage = round((score.marks / 10) * 100, 2)  # Adjust total marks accordingly
            result = 'Pass' if score.marks >= 8 else 'Retraining' if score.marks >= 5 else 'Fail'

            data.append({
                "employee_id": score.employee.id,
                "name": score.employee.full_name,
                "marks": score.marks,
                "percentage": percentage,
                "section": score.employee.department if hasattr(score.employee, 'department') else 'N/A',
                "level_name": score.get_level_number() if score.level else '',
                "skill_name": score.skill.skill if score.skill else '',  # Station.skill string
                "result": result,
            })

        serializer = SimpleScoreSerializer(data, many=True)
        return Response(serializer.data)


class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('skill', flat=True).distinct()
        return Response(skills)


class ScoresBySessionView(APIView):
    def get(self, request, session_key):
        scores = Score.objects.filter(test_name=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EmployeeSkillMatrixView(APIView):
    """
    API endpoint to view an employee's current skill matrix levels
    """
    def get(self, request, employee_id):
        try:
            employee = OperatorMaster.objects.get(id=employee_id)
            skill_levels = OperatorLevel.objects.filter(employee=employee).select_related(
                'operation', 'skill_matrix'
            )

            data = []
            for skill_level in skill_levels:
                data.append({
                    'employee_name': employee.full_name,
                    'employee_code': employee.employee_code,
                    'skill_name': skill_level.operation.name,
                    'current_level': skill_level.level,
                    'department': skill_level.skill_matrix.department,
                    'last_updated': skill_level.operation.matrix.updated_on if skill_level.operation.matrix else None
                })

            return Response(data, status=status.HTTP_200_OK)

        except OperatorMaster.DoesNotExist:
            return Response(
                {'error': f'Employee with id {employee_id} not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Error retrieving skill matrix: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RecentSkillMatrixUpdatesView(APIView):
    """
    API endpoint to view recent skill matrix updates from quiz completions
    """
    def get(self, request):
        try:
            # Get recent scores that passed and have skill/level data
            recent_scores = Score.objects.filter(
                passed=True,
                skill__isnull=False,
                level__isnull=False
            ).select_related('employee', 'skill', 'level').order_by('-created_at')[:20]

            data = []
            for score in recent_scores:
                # Get corresponding skill matrix entry
                operation = OperationList.objects.filter(name__iexact=score.skill.skill).first()
                if operation:
                    operator_level = OperatorLevel.objects.filter(
                        employee=score.employee,
                        operation=operation
                    ).first()

                    data.append({
                        'employee_name': score.employee.full_name,
                        'employee_code': score.employee.employee_code,
                        'skill_name': score.skill.skill,
                        'test_level': score.get_level_number(),
                        'current_skill_level': operator_level.level if operator_level else 0,
                        'test_date': score.created_at,
                        'test_percentage': score.percentage,
                        'test_name': score.test_name
                    })

            return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Error retrieving recent updates: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SkillMatrixView(APIView):
    """
    API view to get skill matrix data in proper format:
    - Employees as rows
    - Skills as columns
    - Level numbers (1, 2, 3) in cells
    """
    def get(self, request, department=None):
        try:
            # Get all employees
            employees = OperatorMaster.objects.all().order_by('full_name')

            # Get skills based on department filter
            if department:
                skill_matrix = SkillMatrix.objects.filter(department__icontains=department).first()
                if skill_matrix:
                    operations = OperationList.objects.filter(matrix=skill_matrix).order_by('name')
                else:
                    operations = OperationList.objects.none()
            else:
                # Get all operations across all departments
                operations = OperationList.objects.all().order_by('matrix__department', 'name')

            # Build skill matrix data
            matrix_data = {
                'departments': [],
                'employees': [],
                'skills': [],
                'matrix': {}
            }

            # Get departments
            departments = SkillMatrix.objects.all().order_by('department')
            for dept in departments:
                dept_operations = OperationList.objects.filter(matrix=dept).order_by('name')
                matrix_data['departments'].append({
                    'name': dept.department,
                    'skills': [op.name for op in dept_operations]
                })

            # Get all skills for columns
            for operation in operations:
                matrix_data['skills'].append({
                    'name': operation.name,
                    'department': operation.matrix.department if operation.matrix else 'Unknown'
                })

            # Get employee data and their skill levels
            for employee in employees:
                employee_data = {
                    'id': employee.id,
                    'name': employee.full_name,
                    'department': employee.department or 'Unknown',
                    'skills': {}
                }

                # Get skill levels for this employee
                operator_levels = OperatorLevel.objects.filter(employee=employee).select_related('operation')

                # Initialize all skills to 0 (no level)
                for operation in operations:
                    employee_data['skills'][operation.name] = 0

                # Fill in actual levels
                for op_level in operator_levels:
                    if op_level.operation.name in employee_data['skills']:
                        employee_data['skills'][op_level.operation.name] = op_level.level

                matrix_data['employees'].append(employee_data)

            return Response(matrix_data)

        except Exception as e:
            print(f"Error in SkillMatrixView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)


# views.py
@api_view(['GET'])
def get_registered_remotes(request):
    paper_id = request.GET.get('paper_id')
    employees = OperatorMaster.objects.filter(question_paper_id=paper_id)
    key_ids = employees.values_list('remote_id', flat=True)  # assuming remote_id field exists
    return Response(list(key_ids))


from rest_framework import viewsets
from .models import CompanyLogo
from .serializers import CompanyLogoSerializer

class CompanyLogoViewSet(viewsets.ModelViewSet):
    queryset = CompanyLogo.objects.all()
    serializer_class = CompanyLogoSerializer



# Refreshment Training


#views.py


from rest_framework import viewsets
from .models import Training_category, Curriculum, CurriculumContent, Trainer_name, Venues, Schedule, OperatorMaster , RescheduleLog,EmployeeAttendance
from .serializers import Training_categorySerializer, CurriculumSerializer, CurriculumContentSerializer, Trainer_nameSerializer, VenuesSerializer, ScheduleSerializer, OperatorMasterSerializer, EmployeeAttendanceSerializer, RescheduleLogSerializer

class Training_categoryViewSet(viewsets.ModelViewSet):
    queryset = Training_category.objects.all()
    serializer_class = Training_categorySerializer

class CurriculumViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumSerializer

    def get_queryset(self):
        queryset = Curriculum.objects.all()
        category_id = self.request.query_params.get('category_id')
        if category_id is not None:
            queryset = queryset.filter(category_id=category_id)
        return queryset

class CurriculumContentViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumContentSerializer

    def get_queryset(self):
        queryset = CurriculumContent.objects.all()
        curriculum_id = self.request.query_params.get('curriculum')
        if curriculum_id is not None:
            queryset = queryset.filter(curriculum_id=curriculum_id)
        return queryset

class Trainer_nameViewSet(viewsets.ModelViewSet):
    queryset = Trainer_name.objects.all()
    serializer_class = Trainer_nameSerializer

class VenueViewSet(viewsets.ModelViewSet):
    queryset = Venues.objects.all()
    serializer_class = VenuesSerializer

class ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action

from .models import EmployeeAttendance, RescheduleLog
from .serializers import EmployeeAttendanceSerializer, RescheduleLogSerializer


class EmployeeAttendanceViewSet(viewsets.ModelViewSet):
    queryset = EmployeeAttendance.objects.all()
    serializer_class = EmployeeAttendanceSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        attendance_instance = serializer.save()

        print(f"Attendance created: status={attendance_instance.status}")

        if attendance_instance.status == 'rescheduled':
            if not (attendance_instance.reschedule_date and attendance_instance.reschedule_time and attendance_instance.reschedule_reason):
                print("Missing reschedule details; skipping RescheduleLog creation")
            else:
                try:
                    RescheduleLog.objects.create(
                        schedule=attendance_instance.schedule,
                        employee=attendance_instance.employee,
                        original_date=attendance_instance.schedule.date,
                        original_time=attendance_instance.schedule.time,
                        new_date=attendance_instance.reschedule_date,
                        new_time=attendance_instance.reschedule_time,
                        reason=attendance_instance.reschedule_reason,
                    )
                    print("RescheduleLog created")
                except Exception as e:
                    print("Error creating RescheduleLog:", e)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def by_schedule(self, request):
        schedule_id = request.query_params.get('schedule_id')
        if not schedule_id:
            return Response({"detail": "schedule_id query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        attendances = self.queryset.filter(schedule_id=schedule_id)
        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)




from rest_framework import viewsets
from rest_framework.response import Response

from .models import RescheduleLog
from .serializers import RescheduleLogSerializer


class RescheduleLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Viewset to retrieve Reschedule Logs (Read-only).
    Optional filtering by schedule or employee.
    """
    queryset = RescheduleLog.objects.all()
    serializer_class = RescheduleLogSerializer

    def list(self, request, *args, **kwargs):
        schedule_id = request.query_params.get('schedule_id')
        employee_id = request.query_params.get('employee_id')

        queryset = self.queryset

        if schedule_id:
            queryset = queryset.filter(schedule_id=schedule_id)

        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


#Master table excel report
from rest_framework import viewsets
from rest_framework.decorators import action
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import OperatorMaster
from datetime import datetime

class OperatorExcelViewSet(viewsets.ViewSet):
    queryset = OperatorMaster.objects.all()
    
    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        queryset = self.queryset
        
        wb = Workbook()
        ws = wb.active
        ws.title = "OPERATOR MASTER"

        # Header setup
        ws.merge_cells('A1:K1')
        ws['A1'] = "Company Name: INDIA JAPAN LIGHTING PVT LTD"
        ws['A1'].font = Font(bold=True, size=10)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A2:K2')
        ws['A2'] = f"Run Date & Time: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10)
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A4:K4')
        ws['A4'] = "OPERATOR MASTER REPORT"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A4'].alignment = Alignment(horizontal="center", vertical="center")

        # Column headers
        headers = [
            'Sr.No.',
            'Employee Code',
            'Full Name',
            'Date of Joining',
            'Pattern Category',
            'Designation',
            'Department',
            'Department Code',
            # Add more headers if needed
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Data rows
        for row_num, operator in enumerate(queryset, 7):
            row_data = [
                operator.sr_no or (row_num - 6),  # Use sr_no if exists, else generate
                operator.employee_code,
                operator.full_name.upper(),
                operator.date_of_join.strftime('%d/%m/%Y') if operator.date_of_join else '',
                operator.employee_pattern_category,
                operator.designation,
                operator.department or '',
                operator.department_code or '',
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(
                    horizontal="center" if col_num in [1, 2, 4] else "left",
                    vertical="center"
                )
                cell.font = Font(size=9)

        # Column widths
        column_widths = {
            'A': 8,   # Sr.No.
            'B': 15,  # Employee Code
            'C': 25,  # Full Name
            'D': 15,  # Date of Joining
            'E': 20,  # Pattern Category
            'F': 20,  # Designation
            'G': 15,  # Department
            'H': 15,  # Department Code
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="OPERATOR_MASTER_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    


 # management review dashboard
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from datetime import datetime
from .models import ManagementReview
from .serializers import (TrainingDataSerializer, DefectsDataSerializer,OperatorsChartSerializer, TrainingPlansChartSerializer,DefectsChartSerializer)

class CurrentMonthTrainingDataView(APIView):
    def get(self, request):
        current_month = timezone.now().replace(day=1)
        try:
            data = ManagementReview.objects.get(
                month_year__year=current_month.year,
                month_year__month=current_month.month
            )
            serializer = TrainingDataSerializer(data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ManagementReview.DoesNotExist:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )

class CurrentMonthDefectsDataView(APIView):
    def get(self, request):
        current_month = timezone.now().replace(day=1)
        try:
            data = ManagementReview.objects.get(
                month_year__year=current_month.year,
                month_year__month=current_month.month
            )
            serializer = DefectsDataSerializer(data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ManagementReview.DoesNotExist:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )

class OperatorsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = OperatorsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TrainingPlansChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = TrainingPlansChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DefectsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = DefectsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ManagementReview
from .serializers import ManagementReviewUploadSerializer

class ManagementReviewUploadView(APIView):
    def post(self, request):
        serializer = ManagementReviewUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            try:
                df = pd.read_excel(excel_file, header=1)
                df.columns = df.columns.str.strip()  # 🧽 Clean trailing/leading spaces from column names

                for _, row in df.iterrows():
                    ManagementReview.objects.update_or_create(
                        month_year=pd.to_datetime(row['Month & Year']).date(),
                        defaults={
                            'new_operators_joined': int(row['New Operators Joined']),
                            'new_operators_trained': int(row['New Operators Trained']),
                            'total_training_plans': int(row['Total Training Plans']),
                            'total_trainings_actual': int(row['Total Trainings Actual']),
                            'total_defects_msil': int(row['Total Defects at MSIL']),
                            'ctq_defects_msil': int(row['CTQ Defects at MSIL']),
                            'total_defects_tier1': int(row['Total Defects at Tier-1']),
                            'ctq_defects_tier1': int(row['CTQ Defects at Tier-1']),  # fixed trailing space issue
                            'total_internal_rejection': int(row['Total Internal Rejection']),
                            'ctq_internal_rejection': int(row['CTQ Internal Rejection']),
                        }
                    )
                return Response({"message": "Data uploaded successfully."}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



from rest_framework import viewsets
from .models import ManagementReview
from .serializers import ManagementReviewSerializer

class ManagementReviewViewSet(viewsets.ModelViewSet):
    queryset = ManagementReview.objects.all().order_by('-month_year')
    serializer_class = ManagementReviewSerializer



from rest_framework import viewsets
from .models import AdvancedManpowerCTQ
from .serializers import NewAdvancedManpowerCTQSerializer

class NewAdvancedManpowerCTQViewSet(viewsets.ModelViewSet):
    queryset = AdvancedManpowerCTQ.objects.all().order_by('-month_year_ctq')
    serializer_class = NewAdvancedManpowerCTQSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import datetime
from .models import AdvancedManpowerCTQ, OperatorRequirement
from .serializers import (
    AdvancedManpowerCTQSerializer,
    OperatorTrendSerializer,
    BufferManpowerTrendSerializer,
    AttritionTrendSerializer,
    AbsenteeTrendSerializer,
    OperatorRequirementSerializer
)

class ManpowerCTQTrendsView(APIView):
    def get(self, request):
        plant = request.query_params.get('plant')
        factory_id = request.query_params.get('factory_id')
        department_id = request.query_params.get('department_id')
        today = datetime.today()

        # ===== CTQ Queryset =====
        ctq_queryset = AdvancedManpowerCTQ.objects.all().order_by('month_year_ctq')
        if plant:
            ctq_queryset = ctq_queryset.filter(plant=plant)
        if factory_id:
            ctq_queryset = ctq_queryset.filter(factory_id=factory_id)
        if department_id:
            ctq_queryset = ctq_queryset.filter(department_id=department_id)

        # Current month data
        current_month_ctq = ctq_queryset.filter(
            month_year_ctq__year=today.year,
            month_year_ctq__month=today.month
        )

        # ===== OperatorRequirement Queryset =====
        operator_queryset = OperatorRequirement.objects.all().order_by('-month')
        if factory_id:
            operator_queryset = operator_queryset.filter(factory_id=factory_id)
        if department_id:
            operator_queryset = operator_queryset.filter(department_id=department_id)

        # ===== Build Flat Response =====
        return Response({
            "current_month": AdvancedManpowerCTQSerializer(current_month_ctq, many=True).data,
            "operator_trend": OperatorTrendSerializer(ctq_queryset, many=True).data,
            "buffer_trend": BufferManpowerTrendSerializer(ctq_queryset, many=True).data,
            "attrition_trend": AttritionTrendSerializer(ctq_queryset, many=True).data,
            "absentee_trend": AbsenteeTrendSerializer(ctq_queryset, many=True).data,
            "operator_requirements": OperatorRequirementSerializer(operator_queryset, many=True).data,
        })
    
from rest_framework import viewsets
from .models import OperatorRequirement
from .serializers import OperatorRequirementSerializer

class OperatorRequirementViewSet(viewsets.ModelViewSet):
    queryset = OperatorRequirement.objects.all().order_by('-month')
    serializer_class = OperatorRequirementSerializer


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import AdvancedManpowerCTQ, Factory, Department, HQ
from .serializers import AdvancedManpowerCTQSerializer

class UploadAdvancedManpowerCTQView(APIView):
    def post(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(excel_file)

            for _, row in df.iterrows():
                factory_name = row['factory_name']
                department_name = row['department_name']

                # Create or get dummy HQ (if not already assigned)
                hq, _ = HQ.objects.get_or_create(name='Default HQ')

                factory, _ = Factory.objects.get_or_create(name=factory_name, defaults={'hq': hq})
                department, _ = Department.objects.get_or_create(name=department_name, factory=factory)

                # Direct creation using model (instead of serializer if name usage)
                AdvancedManpowerCTQ.objects.create(
                    month_year_ctq=row['month_year_ctq'],
                    total_stations_ctq=row['total_stations_ctq'],
                    operator_required_ctq=row['operator_required_ctq'],
                    operator_availability_ctq=row['operator_availability_ctq'],
                    buffer_manpower_required_ctq=row['buffer_manpower_required_ctq'],
                    buffer_manpower_availability_ctq=row['buffer_manpower_availability_ctq'],
                    attrition_trend_ctq=row['attrition_trend_ctq'],
                    absentee_trend_ctq=row['absentee_trend_ctq'],
                    planned_units_ctq=row['planned_units_ctq'],
                    actual_production_ctq=row['actual_production_ctq'],
                    factory=factory,
                    department=department
                )

            return Response({"message": "Excel data uploaded successfully."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#L2fileuploading


from rest_framework import viewsets
from .models import L2FileUpload
from .serializers import L2FileUploadSerializer

class L2FileUploadViewSet(viewsets.ModelViewSet):
    queryset = L2FileUpload.objects.all()
    serializer_class = L2FileUploadSerializer


import os
from rest_framework.views import APIView
from rest_framework.response import Response
from django.conf import settings

class LogoView(APIView):
    def get(self, request):
        image_path = 'logos/nl_logo.jpg'
        full_url = request.build_absolute_uri(settings.MEDIA_URL + image_path)
        return Response({'logo_url': full_url})
    

# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import FileResponse
from .models import UploadedFile
from .serializers import UploadedFileSerializer
import os

class UploadedFileListView(APIView):
    def get(self, request):
        files = UploadedFile.objects.all().order_by('-uploaded_at')
        serializer = UploadedFileSerializer(files, many=True)
        return Response(serializer.data)

class FileDownloadView(APIView):
    def get(self, request, file_id):
        try:
            uploaded_file = UploadedFile.objects.get(id=file_id)
            file_path = uploaded_file.file.path
            response = FileResponse(open(file_path, 'rb'))
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
        except UploadedFile.DoesNotExist:
            return Response({'error': 'File not found'}, status=404)
        

from rest_framework import generics
from .models import Level2TrainingContent
from .serializers import Level2TrainingContentSerializer
from rest_framework.permissions import IsAuthenticatedOrReadOnly

class Level2TrainingContentListCreateView(generics.ListCreateAPIView):
    queryset = Level2TrainingContent.objects.all()
    serializer_class = Level2TrainingContentSerializer
   

    def get_queryset(self):
        queryset = super().get_queryset()
        topic_id = self.request.query_params.get('topic_id')
        if topic_id:
            queryset = queryset.filter(topic_id=topic_id)
        return queryset

class Level2TrainingContentRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Level2TrainingContent.objects.all()
    serializer_class = Level2TrainingContentSerializer
    


from rest_framework import generics
from .models import Level3TrainingContent
from .serializers import Level3TrainingContentSerializer
from rest_framework.permissions import IsAuthenticatedOrReadOnly

class Level3TrainingContentListCreateView(generics.ListCreateAPIView):
    queryset = Level3TrainingContent.objects.all()
    serializer_class = Level3TrainingContentSerializer


    def get_queryset(self):
        queryset = super().get_queryset()
        topic_id = self.request.query_params.get('topic_id')
        if topic_id:
            queryset = queryset.filter(topic_id=topic_id)
        return queryset

class Level3TrainingContentRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Level3TrainingContent.objects.all()
    serializer_class = Level3TrainingContentSerializer




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Department
from .serializers import DepartmentSerializer

class DepartmentByFactoryView(APIView):
    def get(self, request):
        factory_id = request.query_params.get('factory')
        if not factory_id:
            return Response({"error": "Factory ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        departments = Department.objects.filter(factory_id=factory_id)
        serializer = DepartmentSerializer(departments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from rest_framework import generics
from rest_framework.response import Response
from .models import SkillsList, Station
from .serializers import SkillsListSerializer

class SkillsListView(generics.ListAPIView):
    """
    API endpoint to get skills from Station model for dynamic skill assignment
    """
    def get_queryset(self):
        # Return Station objects instead of SkillsList for dynamic skill assignment
        return Station.objects.all()

    def list(self, request, *args, **kwargs):
        stations = self.get_queryset()
        # Return station data with id and skill name for frontend dropdown
        data = [{'id': station.id, 'skill': station.skill} for station in stations]
        return Response(data)

class LevelsListView(generics.ListAPIView):
    """
    API endpoint to get levels from Level model for dynamic level assignment
    """
    def get_queryset(self):
        return Level.objects.all()

    def list(self, request, *args, **kwargs):
        levels = self.get_queryset()
        # Return level data with id and name for frontend dropdown
        data = [{'id': level.id, 'name': level.name, 'name_display': level.get_name_display()} for level in levels]
        return Response(data)




        
# han chou& shoko chou

from .models import HanContent, HanTrainingContent
from .serializers import HanContentSerializer, HanTrainingContentSerializer


class HanContentViewSet(viewsets.ModelViewSet):
    queryset = HanContent.objects.all()
    serializer_class = HanContentSerializer


class HanTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = HanTrainingContent.objects.all()
    serializer_class = HanTrainingContentSerializer


class HanTrainingContentCreateView(APIView):
    def post(self, request):
        serializer = HanTrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import HanTrainingContent
from .serializers import HanTrainingContentSerializer

class HanTrainingContentByContentID(APIView):
    def get(self, request, han_content_id):
        training_contents = HanTrainingContent.objects.filter(han_content_id=han_content_id)
        serializer = HanTrainingContentSerializer(training_contents, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from .models import ShoContent, ShoTrainingContent
from .serializers import ShoContentSerializer, ShoTrainingContentSerializer


class ShoContentViewSet(viewsets.ModelViewSet):
    queryset = ShoContent.objects.all()
    serializer_class = ShoContentSerializer


class ShoTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ShoTrainingContent.objects.all()
    serializer_class = ShoTrainingContentSerializer


class ShoTrainingContentCreateView(APIView):
    def post(self, request):
        serializer = ShoTrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import HanTrainingContent
from .serializers import HanTrainingContentSerializer

class ShoTrainingContentByContentID(APIView):
    def get(self, request, sho_content_id):  # Make sure this matches the URL parameter name
        training_contents = ShoTrainingContent.objects.filter(sho_content_id=sho_content_id)
        serializer = ShoTrainingContentSerializer(training_contents, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

# end


from rest_framework import generics
from .models import SkillMatrix
from .serializers import SkillMatrixHierarchySerializer

class DepartmentHierarchyView(generics.ListAPIView):
    queryset = SkillMatrix.objects.all()
    serializer_class = SkillMatrixHierarchySerializer


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone

from .models import OperatorMaster, MultiSkilling

class AllEmployeesWithActiveSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '').strip()

        # Filter employees by name (case-insensitive)
        operators = OperatorMaster.objects.filter(full_name__icontains=name_query)

        result = []
        today = timezone.now().date()

        for operator in operators:
            # Get all MultiSkilling skills for this operator
            all_skills = MultiSkilling.objects.filter(employee=operator).select_related(
                'department', 'section', 'operation'
            )

            skills_list = []
            for skill in all_skills:
                # Compute dynamic current_status
                if skill.status == 'scheduled' and skill.date and skill.date <= today:
                    current_status = 'in-progress'
                else:
                    current_status = skill.status

                skills_list.append({
                    "skill_level": skill.skill_level,
                    "date": skill.date,
                    "notes": getattr(skill, "remarks", None),  # use remarks if exist or None
                    "status": skill.status,
                    "current_status": current_status,
                    "department": skill.department.department if skill.department else None,  # FK field
                    "section": skill.section.name if skill.section else None,
                    "operation": skill.operation.name if skill.operation else None,
                })

            result.append({
                "employee_id": operator.id,
                "employee_code": operator.employee_code,
                "full_name": operator.full_name,
                "designation": operator.designation,
                "department": operator.department if operator.department else None,  # string field
                "department_code": operator.department_code,
                "date_of_join": operator.date_of_join,
                "employee_pattern_category": operator.employee_pattern_category,
                "skills": skills_list,
            })

        return Response(result, status=status.HTTP_200_OK)






#emp history card
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView

from .models import (
    OperatorMaster,
    OperatorLevel,
    Score,
    MultiSkilling,
    Schedule,
)

from .serializers import (
    CardEmployeeMasterSerializer,
    OperatorCardSkillSerializer,
    CardScoreSerializer,
    CardMultiSkillingSerializer,
    CardRefreshmentTrainingSerializer,
)

import traceback

class EmployeeCardDetailsView(APIView):
    def get(self, request):
        # Accept both 'card_no' and 'employee_code' parameters for compatibility
        employee_code = request.query_params.get('employee_code')
        if not employee_code:
            return Response({'error': 'employee_code parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = OperatorMaster.objects.get(employee_code=employee_code)
        except OperatorMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(traceback.format_exc())
            return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # Fetch and serialize all related data
        employee_data = CardEmployeeMasterSerializer(employee).data
        operator_skills = OperatorCardSkillSerializer(OperatorLevel.objects.filter(employee=employee), many=True).data
        scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
        multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data
        refreshment_training = CardRefreshmentTrainingSerializer(Schedule.objects.filter(employees=employee), many=True).data

        # Construct full response
        response_data = {
            'employee': employee_data,
            'operator_skills': operator_skills,
            'scores': scores,
            'multi_skilling': multi_skilling,
            'refreshment_training': refreshment_training,
        }

        # Print to console
        print("==== Employee Card Details ====")
        import pprint
        pprint.pprint(response_data)  # pretty-print for readability
        print("================================")

        return Response(response_data)

# ==================== NOTIFICATION API VIEWS ====================

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, BasePermission
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from .models import Notification


class NotificationPermission(BasePermission):
    """
    Custom permission for notifications:
    - Allow read access without authentication (for testing)
    - Require authentication for write operations
    """
    def has_permission(self, request, view):
        # Allow read operations (GET, HEAD, OPTIONS) without authentication
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return True
        # Require authentication for write operations
        return request.user and request.user.is_authenticated
from .serializers import (
    NotificationSerializer, NotificationCreateSerializer,
    NotificationUpdateSerializer, NotificationStatsSerializer
)


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing notifications - no authentication required
    """
    serializer_class = NotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        """Return all notifications - no authentication required"""
        return Notification.objects.all().select_related('recipient', 'operator', 'level', 'training_schedule')

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return NotificationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return NotificationUpdateSerializer
        return NotificationSerializer

    def list(self, request, *args, **kwargs):
        """List notifications with filtering options"""
        queryset = self.get_queryset()

        # Filter by read status
        is_read = request.query_params.get('is_read')
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read.lower() == 'true')

        # Filter by notification type (support multiple types)
        notification_types = request.query_params.getlist('notification_type')
        if notification_types:
            queryset = queryset.filter(notification_type__in=notification_types)

        # Filter by single notification type (backward compatibility)
        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)

        # Filter by priority
        priority = request.query_params.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        # Filter by date range
        days = request.query_params.get('days')
        if days:
            try:
                days_int = int(days)
                since_date = timezone.now() - timedelta(days=days_int)
                queryset = queryset.filter(created_at__gte=since_date)
            except ValueError:
                pass

        # Order by creation date (newest first)
        queryset = queryset.order_by('-created_at')

        # Log the query for debugging
        print(f"📊 Notification query: Found {queryset.count()} notifications")
        if notification_types:
            print(f"🔍 Filtered by types: {notification_types}")

        # Debug: Log first few notifications
        for notif in queryset[:3]:
            print(f"🔍 Notification: ID={notif.id}, Type={notif.notification_type}, Title={notif.title[:30]}...")

        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """Mark a notification as read"""
        notification = self.get_object()
        notification.mark_as_read()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_unread(self, request, pk=None):
        """Mark a notification as unread"""
        notification = self.get_object()
        notification.mark_as_unread()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """Mark all notifications as read for the current user"""
        queryset = self.get_queryset().filter(is_read=False)
        count = queryset.count()

        for notification in queryset:
            notification.mark_as_read()

        return Response({
            'message': f'Marked {count} notifications as read',
            'count': count
        })

    @action(detail=False, methods=['get'])
    def unread(self, request):
        """Get only unread notifications"""
        queryset = self.get_queryset().filter(is_read=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get notification statistics"""
        queryset = self.get_queryset()

        total_count = queryset.count()
        unread_count = queryset.filter(is_read=False).count()
        read_count = total_count - unread_count
        recent_count = queryset.filter(
            created_at__gte=timezone.now() - timedelta(hours=24)
        ).count()

        # Group by type
        by_type = dict(queryset.values('notification_type').annotate(
            count=Count('id')
        ).values_list('notification_type', 'count'))

        # Group by priority
        by_priority = dict(queryset.values('priority').annotate(
            count=Count('id')
        ).values_list('priority', 'count'))

        stats_data = {
            'total_count': total_count,
            'unread_count': unread_count,
            'read_count': read_count,
            'recent_count': recent_count,
            'by_type': by_type,
            'by_priority': by_priority
        }

        serializer = NotificationStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Get recent notifications (last 24 hours)"""
        since_date = timezone.now() - timedelta(hours=24)
        queryset = self.get_queryset().filter(created_at__gte=since_date)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@api_view(['GET'])
def notification_count(request):
    """Get unread notification count for the current user"""
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'},
                       status=status.HTTP_401_UNAUTHORIZED)

    count = Notification.objects.filter(
        Q(recipient=request.user) | Q(recipient_email=request.user.email),
        is_read=False
    ).count()

    return Response({'unread_count': count})


@api_view(['GET'])
def test_notifications(request):
    """Test endpoint to debug notification issues"""
    from .models import Notification
    from .serializers import NotificationSerializer

    notifications = Notification.objects.all()[:10]
    serializer = NotificationSerializer(notifications, many=True)

    return Response({
        'count': notifications.count(),
        'notifications': serializer.data,
        'debug': 'This is a test endpoint'
    })


@api_view(['POST'])
def create_system_notification(request):
    """Create a system-wide notification (admin only)"""
    if not request.user.is_authenticated or request.user.role not in ['admin', 'management']:
        return Response({'error': 'Admin access required'},
                       status=status.HTTP_403_FORBIDDEN)

    serializer = NotificationCreateSerializer(data=request.data)
    if serializer.is_valid():
        # If no specific recipient, send to all admin users
        if not serializer.validated_data.get('recipient') and not serializer.validated_data.get('recipient_email'):
            from django.contrib.auth import get_user_model
            User = get_user_model()
            admin_users = User.objects.filter(role__in=['admin', 'management'])

            notifications = []
            for admin in admin_users:
                notification_data = serializer.validated_data.copy()
                notification_data['recipient'] = admin
                notification = Notification.objects.create(**notification_data)
                notifications.append(notification)

            return Response({
                'message': f'Created {len(notifications)} notifications',
                'count': len(notifications)
            }, status=status.HTTP_201_CREATED)
        else:
            notification = serializer.save()
            response_serializer = NotificationSerializer(notification)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def create_test_notification(request):
    """Create a test notification for testing purposes"""
    try:
        from .models import Notification
        from .serializers import NotificationSerializer
        from django.contrib.auth import get_user_model

        User = get_user_model()

        # Get the current user or first user for testing
        recipient = request.user if request.user.is_authenticated else User.objects.first()

        if not recipient:
            return Response({
                'error': 'No users found in the system'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create a test notification
        notification = Notification.objects.create(
            title="Test Notification",
            message="This is a test notification to verify the system is working.",
            notification_type="system_alert",
            recipient=recipient,
            priority="medium",
            metadata={"test": True}
        )

        serializer = NotificationSerializer(notification)
        return Response({
            'message': 'Test notification created successfully',
            'notification': serializer.data,
            'recipient': f'{recipient.first_name} {recipient.last_name}' if recipient else 'Unknown'
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({
            'error': f'Failed to create test notification: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_employee_notification(request):
    """Manually trigger an employee registration notification for testing"""
    try:
        from .models import OperatorMaster, Notification
        from .serializers import NotificationSerializer
        from django.contrib.auth import get_user_model

        User = get_user_model()

        # Get the current user or first user for testing
        recipient = request.user if request.user.is_authenticated else User.objects.first()

        if not recipient:
            return Response({
                'error': 'No users found in the system'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get the latest employee or create a test one
        latest_employee = OperatorMaster.objects.last()

        if not latest_employee:
            return Response({
                'error': 'No employees found in the system. Please add an employee first.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create an employee registration notification
        notification = Notification.objects.create(
            title="New Employee Registered",
            message=f"New employee {latest_employee.full_name} (ID: {latest_employee.employee_code}) has been registered in the system.",
            notification_type='employee_registration',
            recipient=recipient,
            operator=latest_employee,
            priority='medium',
            metadata={
                'employee_code': latest_employee.employee_code,
                'department': latest_employee.department,
                'designation': latest_employee.designation
            }
        )

        serializer = NotificationSerializer(notification)
        return Response({
            'message': 'Employee registration notification created successfully',
            'notification': serializer.data,
            'employee': latest_employee.full_name
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({
            'error': f'Failed to create employee notification: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_all_notification_types(request):
    """Create test notifications for all 7 notification types"""
    try:
        from .models import Notification, OperatorMaster
        from .serializers import NotificationSerializer
        from django.contrib.auth import get_user_model

        User = get_user_model()

        # Get the current user or first user for testing
        recipient = request.user if request.user.is_authenticated else User.objects.first()

        if not recipient:
            return Response({
                'error': 'No users found in the system'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get latest employee for context
        latest_employee = OperatorMaster.objects.last()
        employee_name = latest_employee.full_name if latest_employee else "Test Employee"

        # Define all 7 notification types
        notification_types = [
            {
                'type': 'employee_registration',
                'title': 'New Employee Registered',
                'message': f'{employee_name} has been registered in the system.',
                'priority': 'medium'
            },
            {
                'type': 'level_exam_completed',
                'title': 'Level Exam Completed',
                'message': f'{employee_name} has completed Level 2 evaluation with status: Pass.',
                'priority': 'high'
            },
            {
                'type': 'training_reschedule',
                'title': 'Training Rescheduled',
                'message': f'Training session for {employee_name} has been rescheduled.',
                'priority': 'medium'
            },
            {
                'type': 'refresher_training_scheduled',
                'title': 'Refresher Training Scheduled',
                'message': f'Refresher training has been scheduled for {employee_name}.',
                'priority': 'medium'
            },
            {
                'type': 'refresher_training_completed',
                'title': 'Refresher Training Completed',
                'message': f'{employee_name} has completed refresher training successfully.',
                'priority': 'medium'
            },
            {
                'type': 'bending_training_added',
                'title': 'Bending Training Added',
                'message': 'New bending training module has been added to the system.',
                'priority': 'low'
            },
            {
                'type': 'system_alert',
                'title': 'System Alert',
                'message': 'System maintenance scheduled for tonight at 2:00 AM.',
                'priority': 'urgent'
            }
        ]

        created_notifications = []

        # Create notifications for all types
        for notif_data in notification_types:
            notification = Notification.objects.create(
                title=notif_data['title'],
                message=notif_data['message'],
                notification_type=notif_data['type'],
                recipient=recipient,
                operator=latest_employee if latest_employee else None,
                priority=notif_data['priority'],
                metadata={
                    'test': True,
                    'batch_created': True,
                    'notification_type': notif_data['type']
                }
            )
            created_notifications.append(notification)

        return Response({
            'message': f'Successfully created {len(created_notifications)} notifications for all 7 types',
            'notification_types': [n.notification_type for n in created_notifications],
            'recipient': f'{recipient.first_name} {recipient.last_name}' if recipient else 'Unknown',
            'count': len(created_notifications)
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({
            'error': f'Failed to create notifications: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_all_notifications(request):
    """Delete all notifications for the current user"""
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'},
                       status=status.HTTP_401_UNAUTHORIZED)

    try:
        from .models import Notification
        from django.db.models import Q

        # Delete all notifications for the current user
        deleted_count = Notification.objects.filter(
            Q(recipient=request.user) | Q(recipient_email=request.user.email)
        ).delete()[0]

        return Response({
            'message': f'Deleted {deleted_count} notifications',
            'count': deleted_count
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            'error': f'Failed to delete notifications: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



from rest_framework import generics, status
from rest_framework.response import Response
from .models import UserInfo
from .serializers import UserInfoSerializer

class UserInfoListCreateView(generics.ListCreateAPIView):
    queryset = UserInfo.objects.all()
    serializer_class = UserInfoSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        try:
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            print("❌ Validation or save error:", str(e))
            print("📝 Request data:", request.data)

            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


from rest_framework import generics, status
from rest_framework.response import Response
from .models import HumanBodyCheck
from .serializers import HumanBodyCheckSerializer

class HumanBodyCheckListCreateView(generics.ListCreateAPIView):
    queryset = HumanBodyCheck.objects.all()
    serializer_class = HumanBodyCheckSerializer
    
    def get_queryset(self):
        temp_id = self.request.query_params.get('temp_id')
        if temp_id:
            return HumanBodyCheck.objects.filter(temp_id=temp_id)
        return HumanBodyCheck.objects.all()
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Calculate overall status
        data = serializer.validated_data
        checks = [
            data.get('color_vision', 'pending'),
            data.get('eye_movement', 'pending'),
            data.get('fingers_functionality', 'pending'),
            data.get('hand_deformity', 'pending'),
            data.get('joint_mobility', 'pending'),
            data.get('hearing', 'pending'),
            data.get('bending_ability', 'pending')
        ]
        
        # Check additional checks if any
        additional_checks = data.get('additional_checks', [])
        for check in additional_checks:
            checks.append(check.get('status', 'pending'))
        
        # Determine overall status
        if 'fail' in checks:
            data['overall_status'] = 'fail'
        elif all(status == 'pass' for status in checks):
            data['overall_status'] = 'pass'
        else:
            data['overall_status'] = 'pending'
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    






#SDC Orientation Feedback ViewSet


from rest_framework import viewsets
from .models import SDCOrientationFeedback
from .serializers import SDCOrientationFeedbackSerializer

class SDCOrientationFeedbackViewSet(viewsets.ModelViewSet):
    queryset = SDCOrientationFeedback.objects.all()
    serializer_class = SDCOrientationFeedbackSerializer







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import UserInfo, HumanBodyCheck
from .serializers import FetchUserInfoSerializer

class PassedUsersWithDetailsView(APIView):
    def get(self, request):
        name = request.GET.get('name', '').strip()
        if not name:
            return Response({"error": "Name query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        users = UserInfo.objects.filter(
            first_name__icontains=name
        )

        # Filter users with a passing HumanBodyCheck
        passing_users = [user for user in users if HumanBodyCheck.objects.filter(temp_id=user.temp_id, overall_status='pass').exists()]

        serializer = FetchUserInfoSerializer(passing_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





#SDC Orientation Feedback ViewSet


from rest_framework import viewsets
from .models import SDCOrientationFeedback
from .serializers import SDCOrientationFeedbackSerializer

class SDCOrientationFeedbackViewSet(viewsets.ModelViewSet):
    queryset = SDCOrientationFeedback.objects.all()
    serializer_class = SDCOrientationFeedbackSerializer







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import UserInfo, HumanBodyCheck
from .serializers import FetchUserInfoSerializer

class PassedUsersWithDetailsView(APIView):
    def get(self, request):
        name = request.GET.get('name', '').strip()
        if not name:
            return Response({"error": "Name query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        users = UserInfo.objects.filter(
            first_name__icontains=name
        )

        # Filter users with a passing HumanBodyCheck
        passing_users = [user for user in users if HumanBodyCheck.objects.filter(temp_id=user.temp_id, overall_status='pass').exists()]

        serializer = FetchUserInfoSerializer(passing_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import UserInfo, HumanBodyCheck
from .serializers import FetchUserInfoSerializer

class AllPassedUsersView(APIView):
    def get(self, request):
        # Get temp_ids of all users who passed the human body check
        passed_temp_ids = HumanBodyCheck.objects.filter(overall_status='pass').values_list('temp_id', flat=True).distinct()

        # Filter user info for those passed temp_ids
        passed_users = UserInfo.objects.filter(temp_id__in=passed_temp_ids)

        serializer = FetchUserInfoSerializer(passed_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import UserInfo
from .serializers import ListUserInfoWithBodyCheckSerializer

class UserInfoBodyCheckListView(APIView):
    def get(self, request):
        users = UserInfo.objects.all()
        serializer = ListUserInfoWithBodyCheckSerializer(users, many=True)
        return Response(serializer.data)









from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import SDCOrientationFeedback, UserInfo

class RemainingDepartmentsView(APIView):
    def get(self, request):
        temp_id = request.GET.get('temp_id')

        if not temp_id:
            return Response({"error": "temp_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Check if user exists
        try:
            user = UserInfo.objects.get(temp_id=temp_id)
        except UserInfo.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        # List of all departments
        all_departments = [dept[0] for dept in SDCOrientationFeedback.DEPARTMENT_CHOICES]

        # Departments already submitted for this user
        submitted_departments = SDCOrientationFeedback.objects.filter(user=user).values_list('department', flat=True)

        # Remaining departments
        remaining_departments = list(set(all_departments) - set(submitted_departments))

        return Response({
            "temp_id": temp_id,
            "user": f"{user.first_name} {user.last_name}",
            "remaining_departments": remaining_departments
        }, status=status.HTTP_200_OK)
    


#Employee History PDF
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
import json
import traceback
from datetime import datetime
import os
from .models import OperatorMaster, OperatorLevel, Score, MultiSkilling, Schedule, CompanyLogo

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests
        Accepts both form data and JSON input
        """
        try:
            print("\n=== Received PDF generation request ===")
            
            # 1. Parse input data
            card_no = self._get_card_number(request)
            if not card_no:
                return JsonResponse({'error': 'card_no is required'}, status=400)
            print(f"Processing card_no: {card_no}")

            # 2. Get employee record
            try:
                employee = OperatorMaster.objects.get(employee_code=card_no)
                print(f"Found employee: {employee.full_name}")
            except OperatorMaster.DoesNotExist:
                print(f"Employee not found for card_no: {card_no}")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # 3. Generate PDF content
            print("Generating PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            story = self.create_pdf_content(employee)
            
            # 4. Build PDF document
            print("Building PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="employee_report_{card_no}.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_card_number(self, request):
        """Helper method to extract card_no from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('card_no')
            except json.JSONDecodeError:
                return None
        return request.POST.get('card_no')

    def _format_date(self, date_obj):
        """Helper to format dates safely"""
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
                return date_obj.strftime('%d-%b-%Y')
            except ValueError:
                return 'N/A'
        return date_obj.strftime('%d-%b-%Y') if date_obj else 'N/A'

    def _get_company_logo(self):
        """Get the company logo from database"""
        try:
            logo = CompanyLogo.objects.first()
            if logo and logo.logo:
                return logo.logo.path
        except Exception as e:
            print(f"Error getting company logo: {e}")
        return None

    def create_pdf_content(self, employee):
        """Generate the PDF content structure"""
        styles = getSampleStyleSheet()
        story = []
        
        # Add header with logo and title
        self._add_header(story, styles)
        
        # Add sections only if they have data
        self._add_basic_info(story, styles, employee)
        self._add_operator_skills(story, styles, employee)
        self._add_scores(story, styles, employee)
        self._add_multi_skills(story, styles, employee)
        self._add_refreshment_training(story, styles, employee)
        
        return story

    def _add_header(self, story, styles):
        """Add header with company logo and title using justify-between layout"""
        # Get company logo
        logo_path = self._get_company_logo()
        
        if logo_path and os.path.exists(logo_path):
            try:
                # Create logo image - smaller size
                logo = Image(logo_path, width=1*inch, height=0.7*inch)
                
                # Create title
                title = Paragraph("Employee Comprehensive Report", styles['Title'])
                
                # Create header table with full width and justify-between effect
                header_data = [[title, logo]]
                header_table = Table(header_data, colWidths=[5.5*inch, 1*inch])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),    # Title left(0,0), (-1,-1) aligned
                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'),   # Logo right aligned
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    # Remove any borders or background
                    ('GRID', (0, 0), (-1, -1), 0, colors.white),
                ]))
                story.append(header_table)
            except Exception as e:
                print(f"Error adding logo: {e}")
                # Fallback to title only
                story.append(Paragraph("Employee Comprehensive Report", styles['Title']))
        else:
            # No logo available, just add title
            story.append(Paragraph("Employee Comprehensive Report", styles['Title']))
        
        story.append(Spacer(1, 20))

    def _get_table_style(self):
        """Returns a consistent, professional style for all tables"""
        return TableStyle([
            # Header styling
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),  # Steel blue
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            
            # Data row styling
            ('ALIGN', (0,1), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E0E0E0')),
            
            # Zebra striping
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            
            # Cell padding
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ])

    def _add_basic_info(self, story, styles, employee):
        """Employee basic information from OperatorMaster"""
        # Create section heading
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4682B4'),
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("Basic Information", heading_style))
        story.append(Spacer(1, 8))
        
        basic_data = [
            ["Field", "Value"],
            ["Name", getattr(employee, 'full_name', 'N/A')],
            ["Card No", getattr(employee, 'employee_code', 'N/A')],
            ["Department", getattr(employee, 'department', 'N/A')],
            ["Department Code", getattr(employee, 'department_code', 'N/A')],
            ["Designation", getattr(employee, 'designation', 'N/A')],
            ["Joining Date", self._format_date(getattr(employee, 'date_of_join', None))],
            ["Employee Category", getattr(employee, 'employee_pattern_category', 'N/A')],
            ["Serial Number", str(getattr(employee, 'sr_no', 'N/A'))]
        ]
        
        # Full width table (6.5 inches total)
        basic_table = Table(basic_data, colWidths=[2*inch, 4.5*inch])
        basic_table.setStyle(self._get_table_style())
        story.append(basic_table)
        story.append(Spacer(1, 24))

    def _add_operator_skills(self, story, styles, employee):
        """Operator skills from OperatorLevelTracking"""
        skills = OperatorLevel.objects.filter(employee=employee)
        if not skills.exists():
            return

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4682B4'),
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("Operator Skills", heading_style))
        story.append(Spacer(1, 8))

        skill_data = [["Level", "Department", "Operation", "Date"]]
        for skill in skills:
            skill_data.append([
                str(skill.level),
                skill.skill_matrix.department if hasattr(skill.skill_matrix, "department") else str(skill.skill_matrix_id),
                skill.operation.name if hasattr(skill.operation, "name") else str(skill.operation_id),
                # self._format_date(skill.created_at)
                self._format_date(skill.skill_matrix.updated_on)
            ])

        # Full width table (6.5 inches total)
        skills_table = Table(skill_data, colWidths=[1*inch, 2.2*inch, 2.2*inch, 1.1*inch])
        skills_style = self._get_table_style()
        skills_style.add('ALIGN', (0,1), (0,-1), 'CENTER')  # Center align Level column
        skills_style.add('ALIGN', (3,1), (3,-1), 'CENTER')  # Center align Date column
        skills_table.setStyle(skills_style)
        story.append(skills_table)
        story.append(Spacer(1, 24))

    def _add_scores(self, story, styles, employee):
        """Scores table with proper result formatting"""
        scores = Score.objects.filter(employee=employee)
        if not scores.exists():
            return
            
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4682B4'),
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("Scores and Assessments", heading_style))
        story.append(Spacer(1, 8))
        
        score_data = [["Test", "Marks", "Percentage", "Result", "Date"]]
        for score in scores:
            score_data.append([
                score.test_name or "N/A",
                str(score.marks) if score.marks is not None else "N/A",
                f"{score.percentage}%" if score.percentage is not None else "N/A",
                "Pass" if score.passed else "Fail",
                self._format_date(score.created_at)
            ])
        
        # Full width table (6.5 inches total)
        scores_table = Table(score_data, colWidths=[2.5*inch, 0.8*inch, 1*inch, 0.8*inch, 1.4*inch])
        scores_style = self._get_table_style()
        scores_style.add('ALIGN', (1,1), (2,-1), 'CENTER')  # Center align Marks and Percentage
        scores_style.add('ALIGN', (3,1), (3,-1), 'CENTER')  # Center align Result
        scores_style.add('ALIGN', (4,1), (4,-1), 'CENTER')  # Center align Date
        
        # Add result coloring
        for row in range(1, len(score_data)):
            result = scores[row-1].passed  
            for style_command in self._get_result_badge_style(result, row):
                scores_style.add(*style_command)
        
        scores_table.setStyle(scores_style)
        story.append(scores_table)
        story.append(Spacer(1, 24))

    def _add_multi_skills(self, story, styles, employee):
        """Multi-skilling with proper status formatting"""
        multi_skills = MultiSkilling.objects.filter(employee=employee).select_related('operation', 'section','department')
        if not multi_skills.exists():
            return
            
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4682B4'),
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("Multi-Skilling", heading_style))
        story.append(Spacer(1, 8))
        
        multi_data = [["Skill", "Status", "Department", "Level", "Date"]]
        for skill in multi_skills:
            multi_data.append([
                str(skill.operation.name if skill.operation else skill.skill_level or "N/A"),
                # skill.skill or (skill.operation.name if skill.operation else skill.skill_level or "N/A"),
                skill.status.capitalize() if skill.status else "N/A",
                str(skill.department.department) if skill.department and hasattr(skill.department, 'department') else "N/A",
                skill.skill_level or "N/A",
                self._format_date(skill.date),
            ])
        
        # Full width table (6.5 inches total)
        multi_table = Table(multi_data, colWidths=[2.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 0.9*inch])
        multi_style = self._get_table_style()
        multi_style.add('ALIGN', (1,1), (1,-1), 'CENTER')  # Center align Status
        multi_style.add('ALIGN', (3,1), (3,-1), 'CENTER')  # Center align Level
        multi_style.add('ALIGN', (4,1), (4,-1), 'CENTER')  # Center align Date
        
        # Add status coloring
        for row in range(1, len(multi_data)):
            status = multi_skills[row-1].status
            for style_command in self._get_status_badge_style(status, row):
                multi_style.add(*style_command)
        
        multi_table.setStyle(multi_style)
        story.append(multi_table)
        story.append(Spacer(1, 24))

    def _add_refreshment_training(self, story, styles, employee):
        """Refreshment training with professional layout"""
        trainings = Schedule.objects.filter(employees=employee).select_related(
            'training_category', 'training_name', 'trainer', 'venue'
        )

        if not trainings.exists():
            return

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4682B4'),
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("Refreshment Training", heading_style))
        story.append(Spacer(1, 8))

        training_data = [["Training Topic", "Venue", "Category", "Date", "Status"]]
        for training in trainings:
            training_data.append([
                training.training_name.topic if training.training_name else "N/A",
                training.venue.name if training.venue else "N/A",
                training.training_category.name if training.training_category else "N/A",
                self._format_date(training.date),
                training.status.capitalize() if training.status else "N/A"
            ])

        # Full width table (6.5 inches total)
        training_table = Table(training_data, colWidths=[2.0*inch, 1*inch, 1.5*inch, 0.8*inch, 1.2*inch])
        training_style = self._get_table_style()
        training_style.add('ALIGN', (3,1), (3,-1), 'CENTER')  # Center align Date
        training_style.add('ALIGN', (4,1), (4,-1), 'CENTER')  # Center align Status
        training_table.setStyle(training_style)

        story.append(training_table)
        story.append(Spacer(1, 24))

    def _get_result_badge_style(self, result, row):
        """Returns properly structured style commands for result badges"""
        color = colors.green if result else colors.red
        return [
            ('TEXTCOLOR', (3, row), (3, row), color),
            ('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
        ]

    def _get_status_badge_style(self, status, row):
        """Returns properly structured style commands for status badges"""
        status_colors = {
            'active': colors.green,
            'completed': colors.blue,
            'inprogress': colors.orange,
            'scheduled': colors.purple,
            'inactive': colors.gray,
            'rescheduled': colors.darkblue,
        }
        color = status_colors.get(status.lower(), colors.black)
        return [
            ('TEXTCOLOR', (1, row), (1, row), color),
            ('FONTNAME', (1, row), (1, row), 'Helvetica-Bold')
        ]